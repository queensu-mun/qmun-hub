#!/usr/bin/env python3
"""
Queen's MUN — 2026-27 Collegiate Conference Scouting Schedule
=============================================================
Standalone scouting workbook of collegiate (university-level) Model UN
conferences across North America for the 2026-27 season. Deliberately kept
SEPARATE from the Hub app and the travel budget workbook.

Authoritative source: edit THIS script, not the .xlsx.
Rebuild:  cd ~/qmun-hub/conferences && python3 build_schedule.py

Data compiled 2026-06-10 from official conference sites (see Sources column /
Legend tab). Status codes:
  POSTED  – 2026-27 dates/fees confirmed on the official site
  EST     – inferred from the prior-year cadence; 2026-27 not yet posted
  TBD     – not yet posted / unknown
  VERIFY  – conflicting or uncertain info; confirm before relying on it

Collegiate (university-DELEGATE) conferences only. High-school conferences that
share a host with a collegiate one (HMUN, BosMUN, ILMUN, MUNUC, etc.) are
excluded entirely. Where a host runs both, the Notes column names the HS sibling
only as a disambiguation guard.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# PALETTE (Queen's tricolor on navy, matching the Hub design system)
# ----------------------------------------------------------------------------
NAVY   = "1B2A4A"   # header bands
RED    = "9D1939"   # Queen's red accent
GOLD   = "B89D5E"   # Queen's gold accent
BLUE   = "4B7BBF"   # Queen's blue accent
LIGHT  = "EEF1F6"   # zebra stripe
WHITE  = "FFFFFF"
DARK   = "1B2A4A"

thin = Side(style="thin", color="C9D2E0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(hexc):  return PatternFill("solid", fgColor=hexc)
def f(**kw):     return Font(name="Calibri", **kw)

# Status -> cell fill (quick visual scan of data confidence)
STATUS_FILL = {
    "POSTED": "D6E8D2",   # green-ish: confirmed
    "EST":    "FBEFD0",   # amber: estimated
    "TBD":    "F2D9D9",   # red-ish: unknown
    "VERIFY": "F2D9D9",
}
# Tier -> accent dot color
TIER_FILL = {
    "Top":      RED,
    "Upper-mid": GOLD,
    "Mid":      BLUE,
    "Regional": "9AA6B8",
}

# ----------------------------------------------------------------------------
# THE SCHEDULE DATA
# order = chronological sort key across the Sept 2026 -> Apr 2027 season;
# TBD/undated conferences sort to the bottom (order >= 900).
# tier bucket is one of: Top / Upper-mid / Mid / Regional
# ----------------------------------------------------------------------------
COLS = ["#", "Conference", "Full Name", "Host", "Location", "2026-27 Dates",
        "Status", "Format & Delegation Size", "Tier", "Delegate Fee (per del.)",
        "Registration Window", "Website", "Notes"]

# Each row: (order, region, conf, full, host, loc, dates, status, fmt, tier, fee, reg, web, notes)
ROWS = [
 (10, "US South", "MUNE", "Model United Nations at Emory", "Emory University",
  "Atlanta, GA", "Sep 24-27 OR Oct 1-4, 2026", "VERIFY",
  "Crisis + GA + signature 24-hr committee", "Upper-mid",
  "$60 early / $80 reg / $100 late (USD)",
  "Early closes Jul 14; Reg Sep 15; Late Sep 29",
  "munemory.org", "Home page and reg page list conflicting dates — confirm with secretariat."),

 (20, "US Northeast", "BarMUN", "Boston Area Review Model United Nations", "Boston University",
  "Boston, MA", "Oct 1-4, 2026", "POSTED",
  "~18 committees, GA + crisis. Small=1-11 / Large=12+ (6+ for delegation award)", "Mid",
  "$95 early / $105 reg / $115 late (USD)",
  "Early opens Mar 16; Reg May 10; Late Jul 26 (closes Sep 1)",
  "barmun.org", "BU's collegiate conf (NOT BosMUN, which is high school). Accessible early-season target."),

 (30, "US Northeast", "UPMUNC", "U. Penn Model United Nations Conference", "U. of Pennsylvania",
  "Philadelphia, PA", "Oct 15-18, 2026", "POSTED",
  "College-only. GA + specialized + crisis, competitive", "Top",
  "TBD (2026 not yet posted)",
  "\"Opening soon\" as of Jun 2026",
  "upmunc.org", "Top-tier fall flagship. 'PennMUNC' is just UPMUNC; ILMUNC is Penn's HS conf."),

 (40, "Canada", "CIAC", "Cornell International Affairs Conference", "Cornell University",
  "Ithaca, NY", "Oct 22-25, 2026", "POSTED",
  "GA + specialized + crisis", "Upper-mid",
  "~$80 early / $90 reg; DELEGATES 11+ FREE (USD)",
  "Early ~through Jul 2026",
  "ciaconline.org", "Closest US conf to Kingston (~3.5h). In the budget slate. CIAC = Cornell, not Carleton."),

 (50, "US West", "TrojanMUN", "Trojan Model United Nations", "U. of Southern California",
  "Los Angeles, CA", "Oct 22-25, 2026", "POSTED",
  "13 committees: 7 crisis, 2 GA, 4 specialized. Single + double", "Top",
  "$95 early / $100 reg / $110 late (USD)",
  "Early Apr 8-Jun 19; Reg to Sep 18; Late to Oct 9",
  "trojanmun.org", "#1 Fall West Coast (BestDelegate). USC's SCMUN is the HS conf."),

 (60, "US Northeast", "NCSC", "National Collegiate Security Conference", "Georgetown University",
  "Washington, DC", "Nov 5-8, 2026", "POSTED",
  "Security / crisis-focused, realistic committees", "Top",
  "TBD (2026 not yet posted)",
  "TBD; ncscreg@modelun.org",
  "ncsc.modelun.org", "NCSC LIV. Fall security-circuit anchor (53+ yrs). In the budget slate."),

 (70, "National", "NMUN-DC", "National Model UN — Washington DC", "NCCA / NMUN (UNITAR)",
  "Arlington, VA", "Nov 6-8, 2026", "POSTED",
  "Substantive GA/ECOSOC. Single or double (max 2 per committee). NOT crisis", "Mid",
  "$100/del + $195 delegation (USD)",
  "Opens Apr 15, 2026; all fees due Oct 1",
  "nmun.org", "Smaller fall companion to NMUN-NY. Lower-cost season opener."),

 (80, "US South", "UNCMUNC", "UNC Model United Nations Conference", "U. of North Carolina",
  "Chapel Hill, NC", "Nov 19-22, 2026", "POSTED",
  "GA + specialized + crisis", "Mid",
  "~$105/del (USD)",
  "TBD",
  "—", "In the budget slate (flight from Kingston). No host hotel block."),

 (90, "US West", "NWMUN-Seattle", "Northwest Model UN — Seattle", "NW Assoc. for Global Affairs",
  "Seattle, WA", "Nov 20-22, 2026", "POSTED",
  "Substantive GA-style UN simulation (not crisis-heavy)", "Regional",
  "TBD",
  "TBD",
  "seattle.nwmun.org", "Premier Pacific NW collegiate conf. Far from Kingston."),

 (100, "US Midwest", "AMUN", "American Model United Nations Intl.", "AMUN (independent non-profit)",
  "Chicago, IL", "Nov 21-24, 2026", "POSTED",
  "Substantive GA + Plenary, UNSC, ICJ, CSW. Min 4 del; max 2/committee. NOT crisis", "Mid",
  "$80/del + $120 delegation, flat (USD)",
  "Intl groups must finish by Aug 15",
  "amun.org", "Realism-focused, deliberately non-competitive (no award circuit). 1,400+ attendees."),

 (110, "Canada", "McMUN", "McGill Model United Nations Assembly", "McGill University (IRSAM)",
  "Montreal, QC", "Jan 28-31, 2027", "POSTED",
  "~38 committees, ~1,600 del. GA + specialized + crisis, mostly single", "Top",
  "~$135/del + $80 delegation (2026 fig; incl $20 social)",
  "Priority ~Jun 2026; Regular ~Sep-Dec 2026",
  "mcmun.org", "Largest collegiate MUN in Canada, 3rd in N.A. Flagship target. Sheraton venue. In budget slate."),

 (130, "US West", "SUMUN", "Stanford University Model United Nations", "Stanford University",
  "Stanford, CA", "~early Feb 2027", "EST",
  "Emerging (founded 2024); format TBD", "Mid",
  "$80 early / $90 reg / $100 late (USD)",
  "Early by Dec 31; Reg by Jan 30",
  "stanfordmodelun.org", "Stanford's COLLEGIATE conf. SMUNC is Stanford's HS conf — don't confuse. Financial aid offered."),

 (140, "US South", "Krewe de MUN", "Krewe de MUN", "Tulane University",
  "New Orleans, LA", "~early Feb 2027", "EST",
  "GA + specialized / surprise committees", "Mid",
  "TBD",
  "TBD",
  "events.tulane.edu", "Tulane's collegiate conf (timed near Mardi Gras). No 'TUMUNC' collegiate exists."),

 (150, "US Midwest", "BUCKMUN", "Buckeye Model United Nations", "Ohio State University",
  "Columbus, OH", "Feb 4-7, 2027", "EST",
  "Inaugural edition — committees TBD", "Mid",
  "~$75/del (estimated, inaugural)",
  "TBD — confirm with organizers",
  "—", "INAUGURAL conference, fees estimated. In the budget slate."),

 (160, "US Northeast", "HNMUN", "Harvard National Model United Nations", "Harvard University (IRC)",
  "Boston, MA", "~Feb 11-14, 2027", "EST",
  "Large multi-committee: GA + ECOSOC + regional + crisis. Single & double", "Top",
  "$130/del (2026 fig; 2027 TBD)",
  "Opens ~fall 2026 (2027 TBD)",
  "hnmun.org", "Oldest, largest, most prestigious collegiate conf. Boston Feb 2027 confirmed; exact dates TBD."),

 (170, "Canada", "NAMUN", "North American Model United Nations", "U. of Toronto",
  "Toronto, ON", "~mid-late Feb 2027", "EST",
  "GA + specialized + crisis", "Upper-mid",
  "TBD",
  "TBD (2027 reg not yet live)",
  "namun.org", "Canada's longest-running university conf (42nd ed.). Strong Toronto-circuit target alongside McMUN."),

 (180, "US South", "TechMUN", "Georgia Tech Model United Nations", "Georgia Tech (Sam Nunn School)",
  "Atlanta, GA", "~late Feb / early Mar 2027", "EST",
  "Crisis-forward; specifics TBD", "Mid",
  "TBD",
  "TBD",
  "gtmun.gatech.edu", "Georgia Tech's collegiate conf. GTMUN is their HS conf — don't confuse. Growing."),

 (190, "US West", "UCBMUN", "UC Berkeley Model United Nations", "UC Berkeley",
  "San Francisco, CA", "~early-mid March 2027", "EST",
  "GA + specialized + crisis (UNSC, DISEC)", "Top",
  "$95 early / $105 reg / $115 late (2025-26 fig, USD)",
  "Early Aug-Oct; Reg to Jan; Late to Feb (prior yr)",
  "ucbmun.com", "West Coast's oldest premier spring conf. 'BUCS' = Brown, not Berkeley. BMUN is Berkeley's HS conf."),

 (200, "Canada", "ConMUN", "Concordia Model United Nations", "Concordia University",
  "Montreal, QC", "~mid-March 2027", "EST",
  "6 crisis + 2 specialized + 2 GA, ~300 del. Ages 17-25", "Mid",
  "~$125/del (2026 fig; 2027 not posted)",
  "~Sep 2026 (2027 portal TBD)",
  "conmun.org", "Growing Montreal conf. Good secondary Montreal stop after McMUN. In the budget slate."),

 (210, "National", "NMUN-NY (A)", "National Model UN — New York, Session A", "NCCA / NMUN (UNITAR)",
  "New York, NY", "Mar 21-25, 2027", "POSTED",
  "Substantive GA/ECOSOC, ~16 committees. Single or double (max 2). Closing in UN GA Hall", "Top",
  "$150/del (to Feb 15) + $200 delegation (USD)",
  "Opens Sep 1, 2026; priority Oct 1; pay by Feb 15",
  "nmun.org", "World's largest MUN (5,000+ del, 250+ unis, 58% intl). Substantive prestige track, not head-to-head crisis."),

 (220, "US Midwest", "ChoMUN", "Chicago Model United Nations", "U. of Chicago",
  "Chicago, IL", "Mar 25-28, 2027", "POSTED",
  "~28 committees: GA + crisis + specialized/hybrid. Crisis-heavy. Small<=12 / Large 13+", "Top",
  "$95 priority / $110 reg / $120 late (USD)",
  "Priority Jun 1-Sep 6; Reg to Jan 10; Late to Mar 1",
  "chomun.org", "N.A.'s largest collegiate CRISIS conf (ChoMUN XXX). First-time/legacy discounts. In budget slate. MUNUC is UChicago's HS conf."),

 (230, "US South", "VICS", "Virginia International Crisis Simulation", "U. of Virginia",
  "Charlottesville, VA", "~late March 2027", "EST",
  "All-crisis with GA/specialized", "Top",
  "~$105/del (in budget slate)",
  "TBD (VICS 31 not yet posted)",
  "iro-vics.org", "Top crisis-circuit conf. In the budget slate. VICS = UVA, not Vanderbilt."),

 (240, "National", "NMUN-NY (B)", "National Model UN — New York, Session B", "NCCA / NMUN (UNITAR)",
  "New York, NY", "Mar 29-Apr 2, 2027", "POSTED",
  "Same as Session A — substantive GA/ECOSOC", "Top",
  "$150/del (to Feb 15) + $200 delegation (USD)",
  "Opens Sep 1, 2026; priority Oct 1; pay by Feb 15",
  "nmun.org", "Second NY session, same week-after. Pick A or B."),

 (250, "US Northeast", "PDI", "Princeton Diplomatic Invitational", "Princeton University (IRC)",
  "Princeton, NJ", "~early April 2027", "EST",
  "All-crisis, college-only. Fast-paced single/double crisis rooms", "Top",
  "TBD",
  "TBD; @pdi.26",
  "princetonirc.org/pdihome", "Formerly PICSim — renamed PDI. Top crisis invitational (spring)."),

 (260, "US West", "SGMUN", "Sun God Model United Nations", "UC San Diego",
  "San Diego, CA", "~mid-late April 2027", "EST",
  "Emerging (2-3 yrs old); format TBD", "Mid",
  "TBD",
  "TBD ('Coming Soon')",
  "sgmun.org", "UCSD's collegiate conf. TritonMUN is their HS conf."),

 (270, "US West", "LAMUN", "Los Angeles Model United Nations", "UCLA",
  "Los Angeles, CA", "~late April 2027", "EST",
  "16 committees: 3 GA, 3 specialized, 10 crisis (crisis-heavy)", "Top",
  "$100 early / $105 reg / $110 late (2025-26, USD)",
  "Early Oct-Jan; Reg to Mar; Late to Apr (prior yr)",
  "losangelesmun.org", "Widely regarded #1 West Coast spring conf. UCLA's BruinMUN is the HS conf."),

 # ===== ADDED 2026-06-10 (round 2): BestDelegate + mymun/AllAmericanMUN sweep =====
 (14, "US South", "AIRMUNC", "Alabama IR Model United Nations Conf.", "U. of Alabama (Alabama IRC)",
  "Tuscaloosa, AL", "Oct 1-4, 2026", "POSTED",
  "Ranked collegiate; GA + crisis", "Mid",
  "TBD",
  "TBD",
  "airmunc.org", "First collegiate conf in Alabama. ALMUN is the IRC's HS conf. Site blocks automated fetch."),

 (24, "US Northeast", "FUNMUNC", "Fordham U. National Model UN Conf.", "Fordham University",
  "New York, NY", "Oct 8-11, 2026", "POSTED",
  "GA + crisis + specialized", "Mid",
  "TBD",
  "TBD",
  "funmun.org", "INAUGURAL (FUNMUNC I). Fordham's college travel team currently #1 in N.A."),

 (26, "US South", "BirdMUNC", "Blacksburg IR & Diplomacy MUN Conf.", "Virginia Tech",
  "Blacksburg, VA", "Oct 8-11, 2026", "POSTED",
  "Format TBD", "Regional",
  "TBD",
  "TBD; birdmuncsec@gmail.com",
  "birdmunc.com", "BestDelegate honorable mention. VT's VTMUNC is the HS conf."),

 (32, "US South", "305MUN", "305 Model United Nations", "U. of Miami",
  "Coral Gables, FL", "Oct 15-18, 2026", "POSTED",
  "GA + specialized + crisis. Hundreds of delegates", "Mid",
  "TBD",
  "TBD",
  "305mun.org", "Miami's only collegiate conf; top of the Southern circuit."),

 (54, "US West", "MUNdown", "Buff MUNdown", "U. of Colorado Boulder",
  "Boulder, CO", "~early Nov 2026", "EST",
  "Historical + crisis committees; small one-day event", "Regional",
  "TBD",
  "TBD (Google Form)",
  "cumodelunitednations.com/mundown", "One-day fall event; admits all levels. CU also runs a separate HS conf."),

 (74, "US West", "SBIMUN", "Santa Barbara Intercollegiate MUN", "UC Santa Barbara",
  "Santa Barbara, CA", "Nov 12-15, 2026", "POSTED",
  "GA + specialized/crisis hybrid + crisis. ~360 del, ~24 schools", "Upper-mid",
  "TBD",
  "TBD",
  "sbimun.org", "One of the largest West Coast collegiate conferences by attendance."),

 (76, "US South", "TennMUN", "Tennessee Model United Nations", "U. of Tennessee, Knoxville",
  "Knoxville, TN", "Nov 12-15, 2026", "EST",
  "Format TBD", "Regional",
  "TBD",
  "TBD",
  "tennmun.org", "Year unconfirmed (likely 2026). UTK also runs VolMUN (~late Feb)."),

 (105, "US Midwest", "NDMUNC", "Notre Dame Model United Nations Conf.", "U. of Notre Dame",
  "South Bend, IN", "Dec 3-6, 2026", "POSTED",
  "GA + crisis mix; young/growing", "Mid",
  "TBD",
  "TBD; secretariat@ndmunc.org",
  "sites.nd.edu/ndmunc", "NDMUNC III. Migrating to ndmunc.org."),

 (148, "US South", "SunMUN", "Sunshine State Model United Nations", "U. of Florida MUN team",
  "Orlando, FL", "Feb 4-6, 2027", "POSTED",
  "9 committees: GA + specialized + crisis + ad-hoc; smaller", "Mid",
  "TBD (SunMUN V; ~$60 deleg + $65/del prior yr)",
  "TBD",
  "sunmun.org", "SunMUN V (official rendering 'SunMUN'). UF's collegiate conference, held at Sheraton Lake Buena Vista."),

 (152, "US West", "DevilMUN", "Devil Model United Nations", "Arizona State University",
  "Tempe, AZ", "Feb 4-7, 2027", "POSTED",
  "Crisis-heavy (joint crisis + themed); 4-day", "Regional",
  "TBD",
  "TBD",
  "devilmun.info", "DevilMUN II; first/only intl collegiate conf in Arizona. ASU's SunMUN is the HS conf."),

 (168, "Canada", "AIMUN", "Alberta Intercollegiate Model UN", "Rotating (MacEwan / Mount Royal)",
  "Edmonton or Calgary, AB", "~late Feb / early Mar 2027", "EST",
  "GA + regional bodies (NATO, EU, ASEAN, WHO, UNSC). Small-delegation", "Regional",
  "TBD",
  "TBD",
  "macewanmun.org/aimun-conferences", "Alberta's largest post-secondary MUN, but small (tens of delegates/school). Far from Kingston."),

 (174, "US South", "NoleMUN", "Florida State Model United Nations", "Florida State University (WAP)",
  "Tallahassee, FL", "Feb 25-28, 2027", "POSTED",
  "GA + crisis, mid-size", "Upper-mid",
  "TBD (~$60 deleg + $65/del prior yr)",
  "TBD (opens soon)",
  "nolemun.weebly.com", "NoleMUN VI. FSU is a top-10 N.A. program. Live domain is the weebly, not nolemun.org."),

 (175, "US Northeast", "HopMUNC", "Johns Hopkins Model United Nations Conf.", "Johns Hopkins U. (+ SAIS)",
  "Washington, DC", "Feb 25-28, 2027", "POSTED",
  "Small conference (~150 del); mixed committees", "Mid",
  "TBD ('Coming Soon')",
  "TBD",
  "hopmunc.org", "HopMUNC IV. 'Best Small-Sized Conference' (Delegate's Choice 2025). DC + SAIS backing."),

 (182, "Canada", "MACMUN", "McMaster Model United Nations Conf.", "McMaster University",
  "Hamilton, ON", "~early March 2027", "EST",
  "GA + specialized w/ crisis; beginner/intermediate/advanced tiers", "Mid",
  "TBD",
  "TBD",
  "macmun.org", "CAUTION: as of 2025 opened to BOTH high school AND university delegates (combined). Close to Kingston."),

 (184, "Canada", "QMUNi", "Queen's Model UN Invitational", "Queen's University",
  "Kingston, ON", "~early March 2027", "EST",
  "Crisis + GA; ~300+ del, mixed experience", "Upper-mid",
  "TBD (~CAD $70-80/del historically)",
  "TBD",
  "qmuni.ca", "HOME conference (est. 2014). Distinct from the QMUN Hub project. Confirm details with the team directly."),

 (185, "US Midwest", "HoosierMUN", "Indiana University Model UN Conf.", "Indiana University Bloomington",
  "Bloomington, IN", "~early March 2027", "EST",
  "GA + crisis mix; IU-themed committees", "Mid",
  "TBD (~$189/nt hotel prior yr)",
  "TBD; secgen@hoosiermun.com",
  "hoosiermun.com", "HoosierMUN III. Do NOT confuse with IndianaMUNC, IU's HS conf."),

 (186, "US Northeast", "PUNC", "Pennsylvania United Nations Conference", "Penn State University (PSIADA)",
  "State College, PA", "~early-mid March 2027", "EST",
  "Entirely crisis-based; 2:1 delegate-staffer ratio", "Upper-mid",
  "TBD",
  "TBD",
  "punc.psiada.org", "All-crisis. PSIADA runs a separate HS conf. Not related to UPMUNC at Penn."),

 (245, "US South", "&MUN", "William & Mary Model United Nations", "College of William & Mary (IR Club)",
  "Williamsburg, VA", "~early April 2027", "EST",
  "GA + specialized + crisis", "Upper-mid",
  "TBD",
  "TBD",
  "andmun.org", "&MUN XV. Established. WMHSMUN / WMIDMUN are the same club's HS/MS conferences."),

 (248, "US Northeast", "NYUMUNC", "NYU Model United Nations Conference", "New York University",
  "New York, NY", "Spring 2027 — VERIFY", "VERIFY",
  "Crisis-driven (GA + general/specialized/joint crisis)", "Upper-mid",
  "TBD (tiered early/regular)",
  "~Nov-Mar window (prior yr)",
  "nyumunc.net", "Agents disagreed on timing (Jan/Feb vs April) — confirm. NYUMUNC XVII. EmpireMUNC is NYU's HS conf."),

 (255, "US West", "MUNFW", "Model United Nations of the Far West", "MUNFW consortium (CSU-heavy)",
  "San Francisco, CA", "~April 2027", "EST",
  "Single shared consortium simulation (not competitive/ranked)", "Regional",
  "TBD",
  "TBD",
  "munfw.org", "Est. 1951. Consortium sim, community-college/CSU heavy. Not a typical competitive-circuit add."),

 # ----- Undated / TBD / status-uncertain (sort to bottom) -----
 (900, "Canada", "Carleton Model NATO", "Carleton Model NATO Conference", "Carleton University (NPSIA)",
  "Ottawa, ON", "Winter — TBD", "TBD",
  "NATO simulation (security/transatlantic), ~100 students. Specialized format", "Regional",
  "TBD",
  "TBD",
  "carleton.ca/mnato", "Niche security sim, close to Kingston. NOT 'CIAC' (that's Cornell). Carleton also runs in-house collegiate confs in Sept & Jan."),

 (910, "Canada", "UBCMUN", "UBC Model United Nations", "U. of British Columbia",
  "Vancouver, BC", "~January 2027", "TBD",
  "Western Canada's largest collegiate MUN; format TBD", "Regional",
  "TBD",
  "TBD",
  "ubcmun.com", "Geographically distant for Queen's — low travel priority. Site was unreachable Jun 2026."),

 (920, "US Northeast", "CMUNNY", "Columbia Model UN New York", "Columbia University (CIRCA)",
  "New York, NY", "Fall 2026 — TBD", "TBD",
  "Undergraduate-only. Mixed committees", "Mid",
  "TBD",
  "TBD",
  "columbiamun.org", "CMUNNY XXI. Columbia's HS conf is CMUNCE — exclude that one."),

 (930, "US Northeast", "FCMUN", "Five College Model United Nations", "Five College Consortium",
  "South Hadley, MA", "Fall/Winter — TBD", "TBD",
  "GA + specialized; novice-friendly (crash course offered)", "Regional",
  "~$40/del (very low; prior yr)",
  "TBD",
  "fcmun.org", "Low-cost, accessible regional option for newer delegates. Hosted at Mount Holyoke."),

 (33, "US South", "ArchMUN", "Arch Model United Nations", "U. of Georgia (UGA MUN)",
  "Athens, GA", "Oct 22-25, 2026", "POSTED",
  "GA + crisis committees", "Mid",
  "TBD; ugamun.conferences@gmail.com",
  "TBD",
  "archmun.com", "ArchMUN IV. UGA's collegiate conf (earlier listed as 'DawgMUN'; archmun.com is the live site). UGA is a top-10 N.A. team."),

 (960, "International", "WorldMUN", "Harvard World Model United Nations", "Harvard IRC (rotating co-host)",
  "Abroad — rotates (2027 TBD)", "2027 host TBD", "TBD",
  "Large GA / international format", "Top",
  "TBD",
  "TBD",
  "worldmun.org", "FLAG: hosted OUTSIDE North America (2026 = Lima, Peru; 2027 city decided via bidding). Harvard-run. Included since Barstool MUN covers it; cut if staying strictly N. American."),

 (950, "US Northeast", "SCSY", "Security Council Simulation at Yale", "Yale University (YIRA)",
  "New Haven, CT", "TBD — status uncertain", "VERIFY",
  "Historically 100% crisis/specialized (no large GA)", "Top",
  "TBD",
  "TBD",
  "scsy-yira.org (offline)", "DORMANT? Last confirmed iteration ~2019 (XLII); site offline. Verify it still runs (YUMUNC may be Yale's current collegiate option)."),

 (905, "Canada", "WESMUN", "Western Model United Nations Conf.", "Western University (UWO)",
  "London, ON", "Winter/early spring — TBD", "TBD",
  "GA + likely crisis; ~200+ delegates", "Upper-mid",
  "TBD",
  "TBD",
  "uwomun.org", "Est. 1997, primarily university students. Western's WHIMUN is the HS conf. Site often down."),

 (906, "US Northeast", "BUCS", "Brown University Crisis Simulation", "Brown University",
  "Providence, RI", "~early March — TBD", "TBD",
  "All-crisis", "Mid",
  "TBD; info@browncrisis.org",
  "TBD",
  "browncrisis.org", "Site flaky; socials active. Distinct from BUSUN (Brown's HS conf). Confirm it's still running."),

 (907, "US Midwest", "WUMUNC", "Washington U. Model UN Conference", "Washington U. in St. Louis (WUIRC)",
  "St. Louis, MO", "Fall/winter — TBD", "TBD",
  "All-crisis, mid-size", "Mid",
  "TBD",
  "TBD",
  "wumunc.org", "WashU's COLLEGIATE crisis conf. Distinct from WUMUNS, their HS symposium."),

 (908, "US Northeast", "RISC", "Rutgers International Security Council", "Rutgers University (RUAIR)",
  "New Brunswick, NJ", "TBD (last Oct 23-26, 2025)", "TBD",
  "Security/crisis-focused", "Regional",
  "TBD",
  "TBD",
  "rutgersisc.org", "Rutgers (NJ), not Rhode Island. Rutgers' RUMUN/RUMUNC are HS conferences."),

 (912, "US South", "PegasusMUN", "Pegasus Model United Nations", "U. of Central Florida",
  "Orlando, FL", "Spring 2027 — TBD", "TBD",
  "New conference; format TBD", "Regional",
  "TBD",
  "TBD",
  "pegasusmun.org", "INAUGURAL was 2026 (PegasusMUN I). UCF's KnightMUN is the HS conf."),
]

SOURCES_NOTE = [
 ("Compiled", "2026-06-10 from official conference websites."),
 ("Status — POSTED", "2026-27 dates/fees confirmed on the official site."),
 ("Status — EST", "Inferred from the prior-year cadence; 2026-27 not yet posted."),
 ("Status — TBD", "Not yet posted / unknown."),
 ("Status — VERIFY", "Conflicting or uncertain info; confirm before relying on it."),
 ("", ""),
 ("Firm 2026-27 dates posted", "BarMUN, UPMUNC, CIAC, TrojanMUN, NCSC, NMUN-DC, UNCMUNC,"),
 ("", "NWMUN-Seattle, AMUN, McMUN, NMUN-NY (A&B), ChoMUN."),
 ("", ""),
 ("Recheck Aug-Sep 2026", "Most spring confs post 2026-27 registration in the fall. Re-verify"),
 ("", "HNMUN exact dates, MUNE date conflict, SCSY status, and all EST/TBD fees."),
 ("", ""),
 ("Relationship to budget", "9 of these (UPMUNC, CIAC, NCSC, UNCMUNC, McMUN, BUCKMUN,"),
 ("", "ConMUN, ChoMUN, VICS) are the costed travel slate in"),
 ("", "Queens_MUN_Budget_2026-27.xlsx. This sheet is the wider scouting universe."),
 ("", ""),
 ("Tier legend", "Top = competitive award-circuit flagship; Upper-mid = strong;"),
 ("", "Mid = solid regional draw; Regional = local/accessible/niche."),
 ("", ""),
 ("Caveat", "Tiers are rough circuit reputation, not an official ranking."),
]

# ----------------------------------------------------------------------------
# BUILD
# ----------------------------------------------------------------------------
wb = Workbook()

# ---------- Sheet 1: Schedule ----------
ws = wb.active
ws.title = "Schedule"
ws.sheet_view.showGridLines = False

# Title band
ws.merge_cells("A1:M1")
c = ws["A1"]
c.value = "QUEEN'S MODEL UN  —  2026-27 COLLEGIATE CONFERENCE SCOUTING SCHEDULE"
c.font = f(bold=True, size=15, color=WHITE)
c.fill = fill(NAVY)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:M2")
c = ws["A2"]
c.value = ("North American university-level conferences, Sept 2026 -> Apr 2027.  "
           "Status: POSTED = confirmed · EST = prior-year estimate · TBD = unposted · VERIFY = confirm.  "
           "Compiled 2026-06-10.")
c.font = f(size=9, italic=True, color="44516A")
c.fill = fill(LIGHT)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[2].height = 18

# Header row
hr = 3
for j, name in enumerate(COLS, start=1):
    cell = ws.cell(row=hr, column=j, value=name)
    cell.font = f(bold=True, size=10, color=WHITE)
    cell.fill = fill(RED if name == "Tier" else NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
ws.row_dimensions[hr].height = 30

# Data rows (sorted chronologically)
data = sorted(ROWS, key=lambda r: r[0])
r = hr + 1
for idx, row in enumerate(data, start=1):
    (_order, region, conf, full, host, loc, dates, status, fmt, tier, fee, reg, web, notes) = row
    values = [idx, conf, full, host, loc, dates, status, fmt, tier, fee, reg, web, notes]
    zebra = LIGHT if idx % 2 == 0 else WHITE
    for j, v in enumerate(values, start=1):
        cell = ws.cell(row=r, column=j, value=v)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True,
                                   indent=1 if j in (2, 3, 8, 13) else 0)
        cell.font = f(size=9, color=DARK)
        cell.fill = fill(zebra)
        col = COLS[j-1]
        if col == "Conference":
            cell.font = f(size=10, bold=True, color=NAVY)
        elif col == "Status":
            cell.fill = fill(STATUS_FILL.get(status, zebra))
            cell.font = f(size=9, bold=True, color=DARK)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col == "Tier":
            cell.fill = fill(TIER_FILL.get(tier, zebra))
            cell.font = f(size=9, bold=True, color=WHITE)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col == "Website":
            if web and web not in ("—",):
                url = web.split(" ")[0]
                cell.value = web
                cell.hyperlink = "https://" + url if not url.startswith("http") else url
                cell.font = f(size=9, color="1155CC", underline="single")
    # row height scales a bit with the longest wrapped cell
    longest = max(len(str(fmt)), len(str(notes)), len(str(reg)))
    ws.row_dimensions[r].height = 30 if longest < 55 else (44 if longest < 95 else 58)
    r += 1

# Column widths
widths = [4, 15, 30, 24, 18, 20, 8, 34, 11, 26, 28, 22, 46]
for j, wdt in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(j)].width = wdt

ws.freeze_panes = "B4"
ws.sheet_view.zoomScale = 100

# ---------- Sheet 2: Legend & Sources ----------
ws2 = wb.create_sheet("Legend & Sources")
ws2.sheet_view.showGridLines = False
ws2.merge_cells("A1:B1")
c = ws2["A1"]; c.value = "LEGEND & SOURCES"
c.font = f(bold=True, size=14, color=WHITE); c.fill = fill(NAVY)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws2.row_dimensions[1].height = 28
rr = 3
for k, v in SOURCES_NOTE:
    a = ws2.cell(row=rr, column=1, value=k); b = ws2.cell(row=rr, column=2, value=v)
    a.font = f(bold=True, size=10, color=NAVY); b.font = f(size=10, color=DARK)
    a.alignment = Alignment(horizontal="left", vertical="top", indent=1)
    b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    if k.startswith("Status") or k in ("Firm 2026-27 dates posted", "Recheck Aug-Sep 2026",
                                        "Relationship to budget", "Tier legend"):
        a.fill = fill(LIGHT)
    rr += 1
ws2.column_dimensions["A"].width = 24
ws2.column_dimensions["B"].width = 66

# ----------------------------------------------------------------------------
OUT = "Queens_MUN_Conference_Schedule_2026-27.xlsx"
wb.save(OUT)
print(f"Wrote {OUT}  —  {len(ROWS)} collegiate conferences across {len({r[1] for r in ROWS})} regions.")
