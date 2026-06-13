from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
import os, math

USD_CAD=1.39; HST=1.13; CONTING=1.10; DMG=30.0; SUBS=180.0; MEGABUS=110.0
MONEY='"$"#,##0'  # Excel number format for live-formula currency cells
def rooms(n): return math.ceil(n/4)
def cars(n): return math.ceil(n/4)
def vld_car(rt,days):
    t=46+37*(days-1); t=min(t,220) if days>=6 else t
    return (t+0.29*min(rt,300)+0.19*max(0,rt-300)+DMG)*HST

# cur: USD or CAD ; mode: DRIVE/FLY/BUS ; fixed: forced headcount or None
CONF={
"UPMUNC":dict(full="UPMUNC – Penn",city="Philadelphia, PA",dates="Oct 15-18, 2026",cur="USD",mode="DRIVE",km=617,days=5,hotel=200,deleg=95,perdel=110,free=None,fixed=None,dh="6.5h",note="Verify dates (usually Nov). Off-block $20/del/nt."),
"CIAC":dict(full="CIAC – Cornell",city="Ithaca, NY",dates="Oct 22-25, 2026",cur="USD",mode="DRIVE",km=297,days=4,hotel=149,deleg=75,perdel=90,free=10,fixed=None,regmult=0.5,dh="3.5h",note="Delegates 11+ FREE. Closest. $149/nt incl breakfast. Reg at 50% (confirmed returning-delegation discount; last attended CIAC XIII/2022)."),
"NCSC":dict(full="NCSC – Georgetown",city="Washington, DC",dates="Nov 5-8, 2026",cur="USD",mode="DRIVE",km=814,days=5,hotel=366,deleg=90,perdel=110,free=None,fixed=None,dh="8h",note="Washington Hilton; off-block $150/school/nt. Priciest hotel."),
"UNCMUNC":dict(full="UNCMUNC – UNC",city="Chapel Hill, NC",dates="Nov 19-22, 2026",cur="USD",mode="FLY",km=1350,days=6,flight=425,hotel=140,deleg=100,perdel=105,free=None,fixed=None,dh="fly",note="14-15h each way = fly. No host hotel block."),
"MCMUN":dict(full="McMUN – McGill",city="Montreal, QC",dates="Jan 28-31, 2027",cur="CAD",mode="BUS",hotel=234,deleg=80,perdel=125,free=None,fixed=20,dh="Megabus 3.5h",note="FIXED 20 del. Per-del EST (2027 table unposted, incl $20 social). Sheraton quad $234."),
"BUCKMUN":dict(full="BUCKMUN – Ohio State",city="Columbus, OH",dates="Feb 4-7, 2027",cur="USD",mode="DRIVE",km=951,days=6,hotel=150,deleg=75,perdel=75,free=None,fixed=None,dh="9h",note="INAUGURAL – fees ESTIMATED. Confirm w/ organizers."),
"CONMUN":dict(full="ConMUN – Concordia",city="Montreal, QC",dates="~Mar 11-14, 2027",cur="CAD",mode="BUS",hotel=229,deleg=100,perdel=125,free=None,fixed=20,dh="Megabus 3.5h",note="FIXED 20 del. 2027 dates TBD (2026 figs). DoubleTree $229/rm."),
"CHOMUN":dict(full="ChoMUN – UChicago",city="Chicago, IL",dates="Mar 25-28, 2027",cur="USD",mode="DRIVE",km=1088,days=6,hotel=201,deleg=50,perdel=110,free=None,fixed=None,dh="10h",note="Palmer House $201/quad. First-time deleg $50. 5% CC surcharge."),
"VICS":dict(full="VICS – UVA",city="Charlottesville, VA",dates="~Mar 25-28, 2027",cur="USD",mode="DRIVE",km=1000,days=6,hotel=189,deleg=85,perdel=105,free=None,fixed=None,dh="11-13h",note="VICS 31 dates TBD. Blocks $169-222; used $189."),
"REVMUNC":dict(full="RevMUNC – GW",city="Washington, DC",dates="Feb 18-21, 2027",cur="USD",mode="DRIVE",km=814,days=5,hotel=230,deleg=60,perdel=70,free=None,fixed=None,dh="8h",note="NEW (founded 2026). Same DC drive as NCSC. Regular-window fee; hotel est. Falls on reading week wk2."),
}
order=["UPMUNC","CIAC","NCSC","UNCMUNC","MCMUN","BUCKMUN","REVMUNC","CONMUN","CHOMUN","VICS"]

# Scenario Builder (tab index 1) has a FIXED layout for its 10-conf menu. These addresses let other
# tabs mirror the live scenario. The SB build block asserts it produces exactly these rows.
# Conference menu order + its row block (rows 5..14), so the Selected Slate can point at each row.
SB_ORDER=["UPMUNC","CIAC","NCSC","UNCMUNC","MCMUN","BUCKMUN","REVMUNC","CONMUN","CHOMUN","VICS"]
_SB_R0=5; SB_ROW={k:_SB_R0+i for i,k in enumerate(SB_ORDER)}; _SB_SUBS=15
_SB_SLATE=17; _SB_OPS=18; _SB_ALLIN=19; _SB_AMS=22; _SB_MEM=23; _SB_BAR=24; _SB_DIN=25; _SB_BASE=26; _SB_FEE=28
_SB_FUND=29; _SB_GAP=30; _SB_BEVEN=31; _SB_FX="'Scenario Builder'!$D$3"  # funding total / gap / break-even / FX cell
SBREF=dict(slate=f"'Scenario Builder'!E{_SB_SLATE}",trips=f"'Scenario Builder'!B{_SB_SLATE}",
           ops=f"'Scenario Builder'!E{_SB_OPS}",allin=f"'Scenario Builder'!E{_SB_ALLIN}",
           ams=f"'Scenario Builder'!E{_SB_AMS}",dues=f"'Scenario Builder'!E{_SB_MEM}",
           bar=f"'Scenario Builder'!E{_SB_BAR}",dinner=f"'Scenario Builder'!E{_SB_DIN}",
           base=f"'Scenario Builder'!E{_SB_BASE}",fee=f"'Scenario Builder'!D{_SB_FEE}",
           feerev=f"'Scenario Builder'!E{_SB_FEE}",fund=f"'Scenario Builder'!E{_SB_FUND}",
           gap=f"'Scenario Builder'!E{_SB_GAP}",beven=f"'Scenario Builder'!E{_SB_BEVEN}",fx=_SB_FX)

def eff_n(k,scn): 
    f=CONF[k]['fixed']; return f if f else scn
def bd(k,n):
    c=CONF[k]; return c['free'] if (c['free'] and n>c['free']) else n
def m(x): return "$"+format(int(round(x)),",")

def parts(k,scn):
    c=CONF[k]; n=eff_n(k,scn)
    reg=(c['deleg']+bd(k,n)*c['perdel'])*c.get('regmult',1.0); hotel=rooms(n)*3*c['hotel']
    base_cad=(reg+hotel)*USD_CAD if c['cur']=="USD" else (reg+hotel)
    if c['mode']=="DRIVE": trans=vld_car(c['km']*2,c['days'])*cars(n)
    elif c['mode']=="FLY": trans=n*c['flight']*USD_CAD+850
    else: trans=n*MEGABUS
    sub=base_cad+trans; tot=sub*CONTING
    return dict(n=n,reg=reg,hotel=hotel,base=base_cad,trans=trans,sub=sub,tot=tot)

def split_at(k,n):
    # Decompose a conference's PRE-contingency cost AT A LITERAL headcount n into its USD-denominated
    # dollars and its CAD-native dollars, so the Scenario Builder's FX cell can scale only the USD
    # part live. tot_cad == (usd*FX + cad) * CONTING. Ignores the 'fixed' override so the builder can
    # price Montreal confs at 16 or 20.
    c=CONF[k]
    reg=(c['deleg']+bd(k,n)*c['perdel'])*c.get('regmult',1.0); hotel=rooms(n)*3*c['hotel']
    if c['cur']=="USD":
        usd=reg+hotel
        if c['mode']=="FLY": usd+=n*c['flight']; cad=850.0      # flights USD, ground CAD
        else: cad=vld_car(c['km']*2,c['days'])*cars(n)          # Communauto CAD-native
    else:
        usd=0.0; cad=reg+hotel+n*MEGABUS                        # Montreal: all CAD
    return usd,cad
def parts_split(k,scn):
    return split_at(k,eff_n(k,scn))
# self-check: the split must reproduce parts() at the budgeted FX
for _k in order:
    for _s in (8,12,20):
        _u,_c=parts_split(_k,_s)
        assert abs((_u*USD_CAD+_c)*CONTING - parts(_k,_s)['tot'])<1e-6, f"split mismatch {_k}@{_s}"

# styles
wb=Workbook()
NAVY=PatternFill("solid",fgColor="1F3864");BLUE=PatternFill("solid",fgColor="2E5496")
LBLUE=PatternFill("solid",fgColor="D9E1F2");GREY=PatternFill("solid",fgColor="F2F2F2")
YEL=PatternFill("solid",fgColor="FFF2CC");GREEN=PatternFill("solid",fgColor="E2EFDA");ORANGE=PatternFill("solid",fgColor="FCE4D6")
white=Font(color="FFFFFF",bold=True);bold=Font(bold=True)
thin=Side(style="thin",color="BFBFBF");border=Border(left=thin,right=thin,top=thin,bottom=thin)
def autos(ws,w):
    for i,x in enumerate(w,1): ws.column_dimensions[get_column_letter(i)].width=x

# ---------- SUMMARY ----------
ws=wb.active; ws.title="Summary"
hdr=[["Queen's Model United Nations – Conference Budget 2026-27","","","","","","","",""],
["Transport: US confs = Communauto Value plan + Long Distance · Montreal confs = Megabus · UNCMUNC = fly. Kingston origin. Sourced June 2026.","","","","","","","",""],
["Cars picked up Wednesday PM for an early-Thursday departure, so rental day count includes the Wednesday night (1 extra day per US drive vs a Thu-Sun hold). Exception: CIAC (Cornell, 3.5h) is a Thursday-morning pickup, no extra day.","","","","","","","",""],
["4 del/room · 4 del/car · 3 nights · USD>CAD 1.39 · +10% contingency · McMUN & ConMUN 20 del (16 selectable in Scenario Builder) · excl. meals & faculty.","","","","","","","",""],
["","","","","","","","",""],
["Conference","City","Dates","Travel","8-del* – CAD","8-del* – USD","12-del* – CAD","12-del* – USD","Key note"]]
for r in hdr: ws.append(r)
g8=g12=0
for k in order:
    c=CONF[k]; p8=parts(k,8); p12=parts(k,12); g8+=p8['tot']; g12+=p12['tot']
    trav={"DRIVE":"Commauto "+c['dh'],"FLY":"Fly (YYZ-RDU)","BUS":c['dh']}[c['mode']]
    fx="  (20 fixed)" if c['fixed'] else ""
    ws.append([c['full'],c['city'],c['dates'],trav,m(p8['tot']),m(p8['tot']/USD_CAD),m(p12['tot']),m(p12['tot']/USD_CAD),(("FIXED 20 DEL. " if c['fixed'] else "")+c['note'])])
ws.append(["Conferences subtotal (10)","","","",m(g8),m(g8/USD_CAD),m(g12),m(g12/USD_CAD),"*US confs at 8 or 12; Montreal confs always 20"])
ws.append(["+ Communauto Value subscriptions","","","",m(SUBS),m(SUBS/USD_CAD),m(SUBS),m(SUBS/USD_CAD),"3 accts x $60/yr ($5/mo plan unlocks Long Distance)"])
ws.append(["PROGRAM TOTAL","","","",m(g8+SUBS),m((g8+SUBS)/USD_CAD),m(g12+SUBS),m((g12+SUBS)/USD_CAD),"+3 x $500 refundable Communauto bonds (not a cost)"])
ws.append(["","","","","","","","",""])
ws.append(["FLAGS: McMUN/ConMUN fees CAD-native, partly prior-year (2027 tables unposted) · BUCKMUN inaugural (estimated) · RevMUNC new conf (fee/hotel est, reading-week wk2) · UPMUNC Oct dates unusual · VICS/ConMUN dates TBD.","","","","","","","",""])
# style
ws.merge_cells("A1:I1");ws["A1"].font=Font(bold=True,size=15,color="FFFFFF");ws["A1"].fill=NAVY;ws.row_dimensions[1].height=24
ws.merge_cells("A2:I2");ws.merge_cells("A3:I3");ws["A2"].font=Font(italic=True,size=9);ws["A3"].font=Font(italic=True,size=9)
for cc in range(1,10):
    x=ws.cell(row=5,column=cc);x.fill=BLUE;x.font=white;x.border=border;x.alignment=Alignment(wrap_text=True,vertical="center")
ws.row_dimensions[5].height=28
nrows=len(order)
for i in range(nrows):  # conf rows 6..6+nrows-1
    rr=6+i
    for cc in range(1,10): ws.cell(row=rr,column=cc).border=border
    ws.cell(row=rr,column=9).alignment=Alignment(wrap_text=True);ws.row_dimensions[rr].height=26
    if CONF[order[i]]['fixed']:
        for cc in range(1,10): ws.cell(row=rr,column=cc).fill=YEL
subrow=6+nrows; subsrow=subrow+1; progrow=subrow+2
for cc in range(1,10): ws.cell(row=subrow,column=cc).fill=LBLUE;ws.cell(row=subrow,column=cc).font=bold
for cc in range(1,10): ws.cell(row=progrow,column=cc).fill=NAVY;ws.cell(row=progrow,column=cc).font=white
ws.cell(row=progrow+2,column=1).font=Font(bold=True,color="C00000")
# live scenario panel (mirrors the Scenario Builder) — the matrix above is the fixed reference menu
ph=progrow+4
panel=[("CURRENT SCENARIO  (live from the Scenario Builder)",None,None),
       ("Conference slate (incl Communauto subscriptions)",SBREF['slate'],MONEY),
       ("+ Operating budget (socials / merch / admin)",SBREF['ops'],MONEY),
       ("ALL-IN TOTAL",SBREF['allin'],MONEY),
       ("Delegate-trips",SBREF['trips'],"0"),
       ("Funding (grant + dues + bar + dinner + per-conf fees)",SBREF['fund'],MONEY),
       ("CLUB FUNDS (gap) — the planned subsidy",SBREF['gap'],MONEY),
       ("Break-even delegate fee (gap = $0)",SBREF['beven'],MONEY)]
for j,(lbl,ref,fmt) in enumerate(panel):
    rr=ph+j; ws.cell(row=rr,column=1,value=lbl)
    ws.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=4)
    if ref is not None:
        cell=ws.cell(row=rr,column=5,value=f"={ref}"); cell.number_format=fmt; cell.font=bold; cell.border=border
    if j==0:
        for cc in range(1,10): ws.cell(row=rr,column=cc).fill=NAVY;ws.cell(row=rr,column=cc).font=white
    elif lbl.startswith("ALL-IN"):
        for cc in range(1,6): ws.cell(row=rr,column=cc).fill=LBLUE;ws.cell(row=rr,column=cc).font=bold
    elif lbl.startswith("CLUB FUNDS"):
        ws.cell(row=rr,column=1).font=Font(bold=True,color="C00000");ws.cell(row=rr,column=5).fill=YEL
ws.cell(row=ph+len(panel)+1,column=1,value="Live: edit the Scenario Builder and this panel follows. The conference matrix above is the fixed reference menu (every conf at 8 and 12 delegates).")
ws.cell(row=ph+len(panel)+1,column=1).font=Font(italic=True,size=9,color="1F3864")
ws.merge_cells(start_row=ph+len(panel)+1,start_column=1,end_row=ph+len(panel)+1,end_column=9)
autos(ws,[24,17,17,17,12,12,12,12,52])

# ---------- DETAIL TABS ----------
def detail(scn):
    ws=wb.create_sheet(f"{scn} Delegates")
    ws.append([f"LINE-ITEM DETAIL – {scn}-delegate scenario (Montreal confs fixed at 20)","","","","",""])
    ws.append(["US confs USD>CAD @1.39. Montreal confs CAD-native (USD col = equiv). Communauto/Megabus native CAD, incl tax.","","","","",""])
    ws.append(["","","","","",""])
    ws.append(["Conference / Line item","Basis","Rate","Qty","USD","CAD"])
    def pair(amt,cur):
        return (m(amt),m(amt*USD_CAD)) if cur=="USD" else (m(amt/USD_CAD),m(amt))
    grand=0
    for k in order:
        c=CONF[k];p=parts(k,scn);n=p['n'];cur=c['cur']
        tag=f"  [FIXED {n} DEL]" if c['fixed'] else ""
        ws.append([f"▼ {c['full']}  ({c['city']}, {c['dates']}){tag}","","","","",""])
        u,cd=pair(c['deleg'],cur); ws.append(["   Delegation fee","flat","$"+str(c['deleg'])+f" {cur}","1",u,cd])
        b=bd(k,n);u,cd=pair(b*c['perdel'],cur);fn=" (11+ free)" if (c['free'] and n>c['free']) else ""
        ws.append(["   Delegate fees"+fn,f"{b} billable","$"+str(c['perdel'])+f" {cur}",str(b),u,cd])
        u,cd=pair(c['hotel']*rooms(n)*3,cur);ws.append(["   Hotel",f"{rooms(n)} rms x 3 nts","$"+str(c['hotel'])+f" {cur}",str(rooms(n)*3),u,cd])
        if c['mode']=="FLY":
            ws.append(["   Flights (RT YYZ)",f"{n} delegates","$"+str(c['flight'])+" USD",str(n),m(n*c['flight']),m(n*c['flight']*USD_CAD)])
            ws.append(["   Ground (KGN<>YYZ + xfer)","flat","—","—","~"+m(850/USD_CAD),m(850)])
        elif c['mode']=="BUS":
            ws.append(["   Megabus RT (Kingston<>Montreal)",f"{n} delegates","$110 CAD",str(n),m(n*MEGABUS/USD_CAD),m(n*MEGABUS)])
        else:
            pc=vld_car(c['km']*2,c['days']);t=46+37*(c['days']-1);t=min(t,220) if c['days']>=6 else t
            km=0.29*min(c['km']*2,300)+0.19*max(0,c['km']*2-300)
            ws.append([f"   Communauto Value+LongDist",f"{cars(n)} cars x {m(pc)}/car","gas incl",str(cars(n)),"—",m(pc*cars(n))])
            ws.append([f"      (per car: ${t:.0f} time + ${km:.0f} km + ${pc/HST-t-km:.0f} dmg, x13% HST)","","","","",""])
        ws.append(["   Subtotal","","","",m(p['sub']/USD_CAD),m(p['sub'])])
        ws.append(["   Contingency 10%","","","",m(p['sub']*0.10/USD_CAD),m(p['sub']*0.10)])
        ws.append([f"   TOTAL – {c['full']} ({c['mode']}, {n} del)","","","",m(p['tot']/USD_CAD),m(p['tot'])])
        ws.append(["","","","","",""])
        grand+=p['tot']
    ws.append(["+ Communauto Value subscriptions (3 accts x $60/yr)","","","",m(SUBS/USD_CAD),m(SUBS)])
    ws.append([f"PROGRAM TOTAL – {scn}-del scenario","","","",m((grand+SUBS)/USD_CAD),m(grand+SUBS)])
    # style
    ws.merge_cells("A1:F1");ws["A1"].font=Font(bold=True,size=12,color="FFFFFF");ws["A1"].fill=NAVY
    ws.merge_cells("A2:F2");ws["A2"].font=Font(italic=True,size=9)
    for cc in range(1,7): x=ws.cell(row=4,column=cc);x.fill=BLUE;x.font=white;x.border=border
    for rr in range(5,ws.max_row+1):
        a=str(ws.cell(row=rr,column=1).value or "")
        if a.startswith("▼"):
            fill=YEL if "[FIXED" in a else LBLUE
            for cc in range(1,7): ws.cell(row=rr,column=cc).fill=fill;ws.cell(row=rr,column=cc).font=bold
        elif a.strip().startswith("TOTAL") or a.startswith("PROGRAM"):
            fill=NAVY if a.startswith("PROGRAM") else GREY
            for cc in range(1,7): ws.cell(row=rr,column=cc).fill=fill
            if a.startswith("PROGRAM"):
                for cc in range(1,7): ws.cell(row=rr,column=cc).font=white
            else: ws.cell(row=rr,column=1).font=bold
        elif "Subtotal" in a or "Contingency" in a or a.strip().startswith("(per car") or a.startswith("+ Comm"):
            ws.cell(row=rr,column=1).font=Font(italic=True,size=9)
    autos(ws,[48,22,16,8,13,13])
detail(8); detail(12)

# ---------- SELECTED SLATE (live mirror of the Scenario Builder) ----------
slate=[("NCSC",8),("CIAC",12),("REVMUNC",8),("VICS",12),("CHOMUN",8),("MCMUN",20),("CONMUN",20)]  # preset (drives the static tabs)
ws=wb.create_sheet("Selected Slate")
ws.append(["SELECTED SLATE — live mirror of the Scenario Builder","","","",""])
ws.append(["Updates automatically when you toggle conferences or change delegate counts on the Scenario Builder tab. Grey rows are switched off; on open it shows the planned 7-conf slate. CAD.","","","",""])
ws.append(["","","","",""])
ws.append(["Conference","Go?","Delegates","Travel","Cost +10% (CAD)"])
r0s=5
for i,k in enumerate(SB_ORDER):
    sbr=SB_ROW[k]; rr=r0s+i
    travel={"DRIVE":"Communauto","FLY":"Fly","BUS":"Megabus"}[CONF[k]['mode']]
    ws.cell(row=rr,column=1,value=CONF[k]['full']+("  [16 or 20]" if CONF[k].get('fixed') else ""))
    ws.cell(row=rr,column=2,value=f"='Scenario Builder'!B{sbr}")
    ws.cell(row=rr,column=3,value=f"=IF('Scenario Builder'!B{sbr}=1,'Scenario Builder'!C{sbr},\"—\")")
    ws.cell(row=rr,column=4,value=travel)
    ws.cell(row=rr,column=5,value=f"='Scenario Builder'!E{sbr}").number_format=MONEY
rNs=r0s+len(SB_ORDER)-1
subrow=rNs+1; totrow=subrow+1; allrow=totrow+1
ws.cell(row=subrow,column=1,value="+ Communauto Value subscriptions (3 accts)")
ws.cell(row=subrow,column=5,value=f"='Scenario Builder'!E{_SB_SUBS}").number_format=MONEY
ws.cell(row=totrow,column=1,value="SLATE TOTAL  (conferences + subscriptions)")
ws.cell(row=totrow,column=3,value=f"='Scenario Builder'!B{_SB_SLATE}")
ws.cell(row=totrow,column=4,value="trips")
ws.cell(row=totrow,column=5,value=f"='Scenario Builder'!E{_SB_SLATE}").number_format=MONEY
ws.cell(row=allrow,column=1,value="ALL-IN TOTAL  (+ operating budget)")
ws.cell(row=allrow,column=5,value=f"='Scenario Builder'!E{_SB_ALLIN}").number_format=MONEY
noterow=allrow+2
ws.cell(row=noterow,column=1,value="Live from the Scenario Builder: flip a conference on/off or change a delegate count there and this tab follows. Montreal trips (Megabus) need no car/account; US trips share 3 Communauto Value accounts.")
# style
ws.merge_cells("A1:E1");ws["A1"].font=Font(bold=True,size=12,color="FFFFFF");ws["A1"].fill=NAVY
ws.merge_cells("A2:E2");ws["A2"].font=Font(italic=True,size=9);ws["A2"].alignment=Alignment(wrap_text=True);ws.row_dimensions[2].height=28
for cc in range(1,6): x=ws.cell(row=4,column=cc);x.fill=BLUE;x.font=white;x.border=border
for rr in range(r0s,rNs+1):
    for cc in range(1,6): ws.cell(row=rr,column=cc).border=border
    ws.cell(row=rr,column=2).alignment=Alignment(horizontal="center");ws.cell(row=rr,column=3).alignment=Alignment(horizontal="center")
# green when in (Go=1), grey when out (Go=0) — keyed on the live local Go column
ws.conditional_formatting.add(f"A{r0s}:E{rNs}",FormulaRule(formula=[f"$B{r0s}=1"],fill=GREEN))
ws.conditional_formatting.add(f"A{r0s}:E{rNs}",FormulaRule(formula=[f"$B{r0s}=0"],fill=PatternFill("solid",fgColor="F2F2F2")))
for cc in range(1,6): ws.cell(row=subrow,column=cc).fill=LBLUE;ws.cell(row=subrow,column=cc).font=bold;ws.cell(row=subrow,column=cc).border=border
for cc in range(1,6): c=ws.cell(row=totrow,column=cc);c.fill=NAVY;c.font=white;c.border=border
for cc in range(1,6): c=ws.cell(row=allrow,column=cc);c.fill=LBLUE;c.font=bold;c.border=border
ws.merge_cells(start_row=noterow,start_column=1,end_row=noterow,end_column=5)
c=ws.cell(row=noterow,column=1);c.font=Font(italic=True,size=9,color="1F3864");c.alignment=Alignment(wrap_text=True,vertical="center");ws.row_dimensions[noterow].height=28
autos(ws,[26,8,12,14,18])

# ---------- ASSUMPTIONS ----------
ws=wb.create_sheet("Assumptions & Sources")
A=[["ASSUMPTIONS & SOURCES","",""],["","",""],
["GLOBAL","",""],
["Exchange rate","USD > CAD","1.39"],
["Rooming / car","delegates per room / car","4 / 4 (8>2 rms, 12>3, 20>5)"],
["Nights","Thu-Sun","3"],["Contingency","on subtotal","10%"],
["Montreal confs","McMUN & ConMUN","FIXED 20 delegates each"],
["Excluded","per scope","meals/per-diem, faculty travel"],
["","",""],
["TRANSPORT – US CONFS: COMMUNAUTO VALUE PLAN + LONG DISTANCE","",""],
["Plan","cheapest unlocking Long Distance","Value $5/mo (12-mo min $60/yr/account)"],
["Long Distance time / km","low season Oct16-Jun14","$46 d1 +$37/day · 29c/km <300, 19c after · gas incl"],
["Cars / accounts","1 driver-member per car","8 del=2 / 12 del=3 accounts"],
["Subscriptions","3 Value accts x $60/yr","$180/yr (in PROGRAM TOTAL)"],
["Bonds","per account, refundable","3 x $500 = $1,500 tied up (not a cost)"],
["Why not Open plan","Rules §9.7 + Plan Perks","$500 limit < single long-trip car AND no Long Distance -> unusable"],
["UNCMUNC","14-15h each way","FLY (YYZ-RDU ~$425 RT USD) + $850 ground"],
["","",""],
["TRANSPORT – MONTREAL CONFS: MEGABUS","",""],
["Route","Kingston <> Montreal direct, ~3.5h","Megabus/FlixBus, Kingston term 1175 John Counter"],
["Fare","RT per delegate, book early/together","$110 CAD (no group discount under 40 pax)"],
["","",""],
["PER-CONFERENCE SOURCES & FEES","",""],
["Conference","Registration source","Hotel"]]
src={"UPMUNC":["upmunc.org/conference-registration","Element/W Philadelphia Downtown"],
"CIAC":["ciaconline.org/register","Cayuga Blu Hotel, Ithaca ($149 incl bkfst)"],
"NCSC":["ncsc.modelun.org/registration","Washington Hilton (Dupont Circle)"],
"UNCMUNC":["uncmunc.org/registration","No block – 3-star Chapel Hill ~$140"],
"MCMUN":["mcmun.org (2027 fee table TBD)","Le Centre Sheraton, quad $234 CAD"],
"BUCKMUN":["buckmun.com (fees TBD – buckmun@ccwaosu.org)","Hyatt Place Columbus/OSU ~$150"],
"CONMUN":["conmun.org/en/registration (2026 figs)","DoubleTree by Hilton Montreal $229 CAD"],
"CHOMUN":["chomun.org/registration","Palmer House Hilton ($201/quad)"],
"VICS":["iro-vics.org / iro-vics.app","Hampton/Draftsman/Fairfield ($169-222)"],
"REVMUNC":["gwuias.org/revmunc (secgen.revmunc@gmail.com)","Foggy Bottom / campus-area est ~$230 (no block posted)"]}
for k in order:
    c=CONF[k];cu=c['cur']
    A.append([f"{c['full']} – deleg ${c['deleg']} / del ${c['perdel']} / hotel ${c['hotel']} ({cu})",src[k][0],src[k][1]])
A+=[["","",""],["HOTEL SOURCING QUALITY","",""],
["Actual published block rate","ChoMUN $201, VICS $189, McMUN $234, ConMUN $229","CIAC $149 = prior-year block"],
["Proxy at known host hotel","NCSC $366 (Georgetown group rate), UPMUNC $200",""],
["Open-market estimate (no block exists)","BUCKMUN $150, UNCMUNC $140, RevMUNC $230 (DC)",""],
["","",""],["FLAGS / TO CONFIRM","",""],
["McMUN 2027 per-del fee","table unposted; used $125 CAD est","finance@mcmun.org"],
["ConMUN 2027 dates+fees","2026 figures used","conmun.org"],
["BUCKMUN inaugural","fees estimated","buckmun@ccwaosu.org"],
["RevMUNC new conf (founded 2026)","$70/del reg + $230 DC hotel est; Feb 18-21 = reading-week wk2","secgen.revmunc@gmail.com"],
["UPMUNC Oct 15-18","unusual (usually Nov)","reconfirm"],
["CIAC delegates 11+ free","12-del pays for 10",""],
["CIAC 50% returning discount","CONFIRMED — last attended CIAC XIII (2022); reg halved in budget","sg@ciaconline.org"],
["Communauto: 1 account per car","credit limit + driver rule","8 del=2, 12 del=3 accounts"]]
for r in A: ws.append(r)
ws.merge_cells("A1:C1");ws["A1"].font=Font(bold=True,size=12,color="FFFFFF");ws["A1"].fill=NAVY;ws["A1"].alignment=Alignment(vertical="center")
heads={"GLOBAL","TRANSPORT – US CONFS: COMMUNAUTO VALUE PLAN + LONG DISTANCE","TRANSPORT – MONTREAL CONFS: MEGABUS","PER-CONFERENCE SOURCES & FEES","HOTEL SOURCING QUALITY","FLAGS / TO CONFIRM"}
for rr in range(1,ws.max_row+1):
    a=str(ws.cell(row=rr,column=1).value or "")
    if a in heads:
        for cc in range(1,4): ws.cell(row=rr,column=cc).fill=LBLUE;ws.cell(row=rr,column=cc).font=bold
    if a=="Conference":
        for cc in range(1,4): ws.cell(row=rr,column=cc).fill=BLUE;ws.cell(row=rr,column=cc).font=white
    if a.startswith("Why not Open"): ws.cell(row=rr,column=1).font=Font(bold=True,color="C00000")
autos(ws,[52,42,44])

# ---------- TRANSPORT OPTIONS ----------
def open_car(rt,d): return (60+55*(d-1)+0.34*max(0,rt-75)+DMG)*HST
def van(rt,toll,n): 
    v=1 if n==8 else 2; return v*4*170+v*(rt*12/100*1.50)+v*toll
tolls={"UPMUNC":40,"CIAC":15,"NCSC":30,"REVMUNC":30,"BUCKMUN":50,"CHOMUN":70,"VICS":25}
ws=wb.create_sheet("Transport Options")
ws.append(["TRANSPORT OPTIONS – US drive confs only (Communauto), total CAD","","","","",""])
ws.append(["A=Communauto OPEN (free plan). B=Rental van(s). C=Communauto VALUE + Long Distance (used in budget; needs $5/mo plan + bond).","","","","",""])
ws.append(["","","","","",""])
for scn in (8,12):
    ws.append([f"— {scn} DELEGATES ({cars(scn)} Commauto cars / {1 if scn==8 else 2} van) —","","","","",""])
    ws.append(["Conference","RT km","A: Open","B: Van","C: Value+LD","Cheapest"])
    for k in ["UPMUNC","CIAC","NCSC","REVMUNC","BUCKMUN","CHOMUN","VICS"]:
        c=CONF[k];rt=c['km']*2
        Aa=open_car(rt,c['days'])*cars(scn);Bb=van(rt,tolls[k],scn);Cc=vld_car(rt,c['days'])*cars(scn)
        best=min([("Open",Aa),("Van",Bb),("Value-LD",Cc)],key=lambda t:t[1])[0]
        ws.append([c['full'],str(rt),m(Aa),m(Bb),m(Cc),best])
    ws.append(["","","","","",""])
ws.append(["READ-OUT","","","","",""])
ws.append(["• Open never wins and can't clear its own $500 credit limit on long trips. Value+LD used in budget.","","","","",""])
ws.append(["• At 8 del a single van is cheapest; at 12 del Value+LD (3 small cars on Long Distance) edges 2 vans.","","","","",""])
ws.append(["• Montreal confs (McMUN/ConMUN) use Megabus $110 CAD RT/delegate – not shown here.","","","","",""])
ws.merge_cells("A1:F1");ws["A1"].font=Font(bold=True,size=12,color="FFFFFF");ws["A1"].fill=NAVY
ws.merge_cells("A2:F2");ws["A2"].font=Font(italic=True,size=9)
for rr in range(1,ws.max_row+1):
    a=str(ws.cell(row=rr,column=1).value or "")
    if a.startswith("—"): 
        for cc in range(1,7): ws.cell(row=rr,column=cc).fill=LBLUE;ws.cell(row=rr,column=cc).font=bold
    if a=="Conference":
        for cc in range(1,7): x=ws.cell(row=rr,column=cc);x.fill=BLUE;x.font=white;x.border=border
    if a=="READ-OUT": ws.cell(row=rr,column=1).font=Font(bold=True,color="C00000")
autos(ws,[24,8,12,12,14,11])


# ---------- PAYMENT TIMELINE ----------
ws=wb.create_sheet("Payment Timeline")
ws.append(["PAYMENT TIMELINE – Selected Slate (7 confs)  ·  today: Jun 10, 2026","","","",""])
ws.append(["Chronological money + action schedule. Amounts = budgeted (regular-window) fees. Register in EARLY/PRIORITY windows to save ~$10-15/del where noted.","","","",""])
ws.append(["","","","",""])
ws.append(["When","Conference","Action / Milestone","Amount due","Notes"])
T=[
("PHASE 1 — REGISTER EARLY (lock rates, pay delegation deposits)","","","",""),
("now → Aug 23, 2026","NCSC","Register (Regular window) + delegation deposit","$90 USD","Priority closed Apr 26; deposit non-refundable"),
("now → Jul 31, 2026","CIAC","Register (Early) + delegation deposit","$75 USD","Early locks $80/del vs $90; delegates 11+ free"),
("Jun 1 → Sep 6, 2026","ChoMUN","Register (Priority) + delegation deposit","$50 USD","First-time-school discount; locks $95/del priority"),
("summer → Sep 2026","McMUN","Register when 2027 portal opens + deposit","$80 CAD","Watch mcmun.org; priority wave cheapest"),
("~Sep 2026","ConMUN","Register (Early) + delegation deposit","$100 CAD","2027 portal TBD; early ~$110/del"),
("Aug 10 → Sep 2, 2026","RevMUNC","Register (Early window) + interest form","$60 USD","Early locks $60/del vs $70; interest form = 10% off"),
("by Sep 1, 2026","NCSC","Apply for NCSCAid (need-based grant)","saves TBD","Must be REGISTERED first. Inaugural fixed pool; email k.lozada@modelun.org for the amount"),
("Aug 31 / Nov 16, 2026","McMUN","Apply for merit invoice discount (priority / regular wave)","saves TBD","Discounts the whole invoice; decisions Sep 15 / Dec 1, then 2 wks to pay. finance@mcmun.org"),
("PHASE 2 — SET UP TRANSPORT (before first trip, Oct 2026)","","","",""),
("by mid-Oct 2026","Communauto","Open 3 Value accounts ($5/mo) + post bonds","$1,500 CAD","Bonds REFUNDABLE; lifts credit limit to $1,000/acct"),
("PHASE 3 — FALL CONFERENCES","","","",""),
("by Oct 7, 2026","NCSC","Final delegate payment due","$880 USD","8 x $110; hard payment deadline"),
("by Oct 9, 2026","CIAC","Late reg closes / balance due","$900 USD","10 x $90 (12 del, 2 free)"),
("Oct 22-25, 2026","CIAC","CONFERENCE — Communauto trip + hotel","~$1,463 CAD","+ hotel $1,341 USD to Cayuga Blu"),
("Nov 5-8, 2026","NCSC","CONFERENCE — Communauto trip + hotel","~$1,770 CAD","+ hotel $2,196 USD to Washington Hilton"),
("PHASE 4 — WINTER / SPRING CONFERENCES","","","",""),
("by ~Dec 25, 2026","McMUN","Hotel block cutoff (Le Centre Sheraton)","$3,510 CAD","5 rms x 3 nts x $234; reconfirm 2027 cutoff date"),
("by ~late Dec 2026","McMUN","Book Megabus (~1 mo out) + final delegate pay","$2,200 + $2,500 CAD","Bus 20x$110; del fees 20x$125 (est)"),
("Jan 10, 2027","ChoMUN","Regular registration closes","—","Late window after = higher per-del fee"),
("Jan 19, 2027","VICS","Delegate fees refundable until this date","—","After: owed in full even if you withdraw"),
("Jan 28-31, 2027","McMUN","CONFERENCE — Megabus + hotel","(paid above)","Sheraton, Montreal"),
("by ~Feb 2027","ConMUN","Book Megabus + final delegate payment","$2,200 + $2,500 CAD","Hotel $3,435 CAD to DoubleTree"),
("Feb 12, 2027","RevMUNC","Late reg closes / balance due","$560 USD","8 x $70; pay before late window steps up to $80"),
("Feb 18-21, 2027","RevMUNC","CONFERENCE — Communauto + hotel","~$1,189 CAD","DC drive (same as NCSC) + hotel $1,380 USD; reading-week wk2"),
("Mar 1, 2027","VICS","Late registration closes","—","VICS 31 dates/windows TBD — verify"),
("Mar 5, 2027","ChoMUN","Payment deadline","$880 USD","8 x $110; +5% if paying by credit card"),
("Mar 11-14, 2027","ConMUN","CONFERENCE — Megabus + hotel","(paid above)","DoubleTree, Montreal (2027 dates TBD)"),
("Mar 25-28, 2027","ChoMUN","CONFERENCE — Communauto + hotel","~$2,315 CAD","+ hotel $1,206 USD to Palmer House"),
("Mar 25-28, 2027","VICS","CONFERENCE — Communauto + hotel + balance","~$3,270 CAD","+ del fees $1,260 USD + hotel $1,701 USD"),
("AFTER SEASON","Communauto","Reclaim 3 x $500 bonds","+$1,500 CAD","Refundable after 1 yr (3 mo notice)"),
]
for r in T: ws.append(r)
ws.merge_cells("A1:E1");ws["A1"].font=Font(bold=True,size=13,color="FFFFFF");ws["A1"].fill=NAVY
ws.merge_cells("A2:E2");ws["A2"].font=Font(italic=True,size=9)
for cc in range(1,6): x=ws.cell(row=4,column=cc);x.fill=BLUE;x.font=white;x.border=border
for rr in range(5,ws.max_row+1):
    a=str(ws.cell(row=rr,column=1).value or "")
    b=str(ws.cell(row=rr,column=3).value or "")
    if a.startswith("PHASE"):
        ws.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=5)
        ws.cell(row=rr,column=1).fill=LBLUE;ws.cell(row=rr,column=1).font=bold
    elif "CONFERENCE" in b:
        for cc in range(1,6): ws.cell(row=rr,column=cc).fill=YEL
        ws.cell(row=rr,column=2).font=bold
    elif a=="AFTER SEASON":
        for cc in range(1,6): ws.cell(row=rr,column=cc).fill=GREEN
    elif b.startswith("Apply for"):
        for cc in range(1,6): ws.cell(row=rr,column=cc).fill=GREEN
        ws.cell(row=rr,column=2).font=bold
    for cc in range(1,6): ws.cell(row=rr,column=cc).border=border
    ws.cell(row=rr,column=5).alignment=Alignment(wrap_text=True)
    ws.cell(row=rr,column=4).font=Font(bold=True)
autos(ws,[20,14,40,18,40])

# ---------- RAIL SCREEN ----------
ws=wb.create_sheet("Rail Screen")
ws.append(["RAIL SCREEN – can we take the train instead of driving/flying?  (excl. McMUN/ConMUN, which use Megabus)","","","","",""])
ws.append(["Verdict: NON-STARTER for all US confs. Kingston has NO Amtrak — every trip = VIA Rail to a gateway + Amtrak cross-border, which forces 1-2 hotel overnights before you reach the US. Sourced June 2026.","","","","",""])
ws.append(["","","","","",""])
ws.append(["THE STRUCTURAL KILLER","","","","",""])
ws.append(["• Only cross-border trains: Maple Leaf (Toronto>NYC) & Adirondack (Montreal>NYC) — both ~12-13h incl. ~2h border stop.","","","","",""])
ws.append(["• Both depart their Canadian gateway in the MORNING and arrive NYC at NIGHT — too late for any same-day southbound leg.","","","","",""])
ws.append(["• Result: a Toronto/Montreal overnight before the US, then a 2nd NYC overnight before going south. No US conf reachable by train in under 2 days.","","","","",""])
ws.append(["","","","","",""])
ws.append(["Conference","City","Current plan","Best train (moving time)","Rail fare ~1-way","Verdict"])
RAIL=[
("CIAC","Ithaca, NY","Drive 3.5h","No rail to city — nearest Amtrak is Syracuse + 60 mi by car","n/a","ABSURD — Ithaca has zero Amtrak service"),
("UPMUNC","Philadelphia, PA","Drive 6.5h","~16h, 2 transfers","US$185-210","Non-starter (1-2 overnights each way)"),
("NCSC","Washington, DC","Drive 8h","~18h, 2 transfers","US$200-230","Non-starter (1-2 overnights each way)"),
("BUCKMUN","Columbus, OH","Drive 9h","No rail to city at all","n/a","IMPOSSIBLE — Columbus has zero Amtrak (largest US metro w/ none)"),
("CHOMUN","Chicago, IL","Drive 10h","~34h, 3 transfers","US$300+","Non-starter (2 hotels + on-train overnight)"),
("VICS","Charlottesville, VA","Drive 11-13h","~22-26h, 2-3 transfers","US$215-250","Non-starter (closest call; drive is also brutal)"),
("UNCMUNC","Durham/Raleigh, NC","Fly $425 RT","~26h, 3 transfers","~US$240 (~$480 RT)","Non-starter — RT fare ~matches flight, +4 hotel nights & ~5x the time"),
]
for r in RAIL: ws.append(r)
ws.append(["","","","","",""])
ws.append(["REFERENCE LEGS (verified 2026)","","","","",""])
ws.append(["VIA Rail Kingston>Toronto","~2h","~C$44-65 econ","Frequent daily corridor service","",""])
ws.append(["VIA Rail Kingston>Montreal","~2.5h","~C$30-70 econ","","",""])
ws.append(["Amtrak Maple Leaf Toronto>NYC","~12-13h","~US$90-150","Dep Toronto 08:20, arr NY Penn ~21:15; ~2h border @ Niagara. Running 2026.","",""])
ws.append(["Amtrak Adirondack Montreal>NYC","~11-12h","~US$70-130","Resumed after 2023-24 suspension; running 2026.","",""])
ws.append(["Amtrak NE Regional NYC>DC / >Philadelphia","3.5h / 1.5h","from US$20 / US$25","Frequent same-day once on the NEC","",""])
ws.append(["Amtrak Lake Shore Limited NYC>Chicago","~19-20h","from US$172","Dep NY 15:40 — hours before Maple Leaf arrives, so no same-day connect","",""])
ws.append(["Amtrak Carolinian NYC/DC>Durham NC","~11-13h / ~7h","from US$86 / US$59","One-seat ride; the NC end is good, the Kingston>NEC end is the problem","",""])
ws.append(["","","","","",""])
ws.append(["TWO NEAR-MISSES (still rejected)","","","","",""])
ws.append(["• UNCMUNC: NC end is great (Carolinian one-seat to Durham), but RT rail ~= the $425 flight and adds ~4 hotel nights + days of transit. Fly stays correct.","","","","",""])
ws.append(["• VICS: only one where rail's downside shrinks — because a 13h drive is itself miserable. Still 2 overnights each way. Not viable for a student team.","","","","",""])
ws.append(["","","","","",""])
ws.append(["NET: nothing changes in the budget. Communauto wins the close US trips, flying wins UNCMUNC, Megabus wins Montreal. Rail beats none of them.","","","","",""])
ws.merge_cells("A1:F1");ws["A1"].font=Font(bold=True,size=12,color="FFFFFF");ws["A1"].fill=NAVY
ws.merge_cells("A2:F2");ws["A2"].font=Font(italic=True,size=9);ws.row_dimensions[2].height=28;ws["A2"].alignment=Alignment(wrap_text=True,vertical="center")
for rr in range(3,ws.max_row+1):
    a=str(ws.cell(row=rr,column=1).value or "")
    if a in ("THE STRUCTURAL KILLER","REFERENCE LEGS (verified 2026)","TWO NEAR-MISSES (still rejected)"):
        for cc in range(1,7): ws.cell(row=rr,column=cc).fill=LBLUE;ws.cell(row=rr,column=cc).font=bold
    elif a=="Conference":
        for cc in range(1,7): x=ws.cell(row=rr,column=cc);x.fill=BLUE;x.font=white;x.border=border
    elif a.startswith("NET:"):
        for cc in range(1,7): ws.cell(row=rr,column=cc).fill=GREEN
        ws.cell(row=rr,column=1).font=Font(bold=True,color="1F3864")
for i in range(len(RAIL)):
    rr=10+i
    for cc in range(1,7): ws.cell(row=rr,column=cc).border=border
    ws.cell(row=rr,column=6).alignment=Alignment(wrap_text=True);ws.cell(row=rr,column=4).alignment=Alignment(wrap_text=True)
    ws.row_dimensions[rr].height=28
    v=str(ws.cell(row=rr,column=6).value or "")
    if v.startswith("ABSURD") or v.startswith("IMPOSSIBLE"):
        ws.cell(row=rr,column=6).font=Font(bold=True,color="C00000")
autos(ws,[14,18,14,40,15,40])

# ---------- DELEGATE FEES (live mirror of the Scenario Builder) ----------
FEES=[100,200,250,300,350,400]
def sm(x):  # signed money
    x=int(round(x)); s="-" if x<0 else ""; return f"{s}${abs(x):,}"
RED=Font(color="C00000")
FEE_ABS=f"'Scenario Builder'!$D${_SB_FEE}"; TRIPS_ABS=f"'Scenario Builder'!$B${_SB_SLATE}"; SLCOST_ABS=f"'Scenario Builder'!$E${_SB_SLATE}"
ws=wb.create_sheet("Delegate Fees")
ws.append(["DELEGATE FEES — live: what each delegate's fee covers, per the current Scenario Builder slate","","","","",""])
ws.append(["NET @ fee = that conference's own Fee /del × delegates − its all-in cost (reg + hotel + travel +10%). Negative (red) = club subsidizes; positive (green) = surplus. Each fee is set per conference on the Scenario Builder. Toggle confs / counts / per-conf fees there and this follows. CAD.","","","","",""])
ws.append(["Break-even /del = the per-delegate fee that exactly covers THAT conference. For reference, a delegate typically pays ~$300-450/conf.","","","","",""])
ws.append(["","","","","",""])
ws.append(["Conference","Go?","Del","Club cost (CAD)","Break-even /del","Net @ fee"])
r0f=6
for i,k in enumerate(SB_ORDER):
    sbr=SB_ROW[k]; rr=r0f+i
    g=f"'Scenario Builder'!B{sbr}"; d=f"'Scenario Builder'!C{sbr}"; cost=f"'Scenario Builder'!D{sbr}"; ec=f"'Scenario Builder'!E{sbr}"; fee=f"'Scenario Builder'!F{sbr}"
    ws.cell(row=rr,column=1,value=CONF[k]['full']+("  [16 or 20]" if CONF[k].get('fixed') else ""))
    ws.cell(row=rr,column=2,value=f"={g}")
    ws.cell(row=rr,column=3,value=f'=IF({g}=1,{d},"—")')
    ws.cell(row=rr,column=4,value=f"={ec}").number_format=MONEY
    ws.cell(row=rr,column=5,value=f'=IF({g}=1,{cost}/{d},"—")').number_format=MONEY
    ws.cell(row=rr,column=6,value=f'=IF({g}=1,{fee}*{d}-{cost},"—")').number_format=MONEY
rNf=r0f+len(SB_ORDER)-1
totf=rNf+1
ws.cell(row=totf,column=1,value="TOTAL (current slate, incl subscriptions)")
ws.cell(row=totf,column=3,value=f"={TRIPS_ABS}")
ws.cell(row=totf,column=4,value=f"={SLCOST_ABS}").number_format=MONEY
ws.cell(row=totf,column=5,value=f'=IF({TRIPS_ABS}=0,"",{SLCOST_ABS}/{TRIPS_ABS})').number_format=MONEY
ws.cell(row=totf,column=6,value=f"='Scenario Builder'!E{_SB_FEE}-{SLCOST_ABS}").number_format=MONEY
# flat-fee sensitivity (live: revenue = fee × live trips, cost = live slate)
sh=totf+2
ws.cell(row=sh,column=1,value="REVENUE vs COST AT A FLAT FEE  (live — whole current slate)")
ws.cell(row=sh+1,column=1,value="Flat fee / delegate");ws.cell(row=sh+1,column=2,value="Total revenue");ws.cell(row=sh+1,column=3,value="Slate cost");ws.cell(row=sh+1,column=4,value="Surplus / (shortfall)");ws.cell(row=sh+1,column=5,value="% of cost covered")
for j,f in enumerate(FEES):
    rr=sh+2+j
    ws.cell(row=rr,column=1,value=f"${f}")
    ws.cell(row=rr,column=2,value=f"={f}*{TRIPS_ABS}").number_format=MONEY
    ws.cell(row=rr,column=3,value=f"={SLCOST_ABS}").number_format=MONEY
    ws.cell(row=rr,column=4,value=f"={f}*{TRIPS_ABS}-{SLCOST_ABS}").number_format=MONEY
    ws.cell(row=rr,column=5,value=f'=IF({SLCOST_ABS}=0,"",{f}*{TRIPS_ABS}/{SLCOST_ABS})').number_format="0%"
blr=sh+2+len(FEES)+1
ws.cell(row=blr,column=1,value="Slate break-even fee (excl operating) →");ws.cell(row=blr,column=4,value=f'=IF({TRIPS_ABS}=0,"",{SLCOST_ABS}/{TRIPS_ABS})').number_format=MONEY
note=blr+1
ws.cell(row=note,column=1,value="This tab covers the conference slate only. The full-program break-even (after operating costs, the AMS grant, dues and fundraising) is the Break-even cell on the Scenario Builder. The cheap-headcount Montreal confs throw off surplus that offsets the expensive flying/long-drive ones.")
# style
ws.merge_cells("A1:F1");ws["A1"].font=Font(bold=True,size=12,color="FFFFFF");ws["A1"].fill=NAVY;ws.row_dimensions[1].height=22
ws.merge_cells("A2:F2");ws.merge_cells("A3:F3");ws["A2"].font=Font(italic=True,size=9);ws["A3"].font=Font(italic=True,size=9)
ws["A2"].alignment=Alignment(wrap_text=True);ws.row_dimensions[2].height=40
hdrrow=5
for cc in range(1,7): x=ws.cell(row=hdrrow,column=cc);x.fill=BLUE;x.font=white;x.border=border;x.alignment=Alignment(wrap_text=True,vertical="center")
for rr in range(r0f,rNf+1):
    for cc in range(1,7): ws.cell(row=rr,column=cc).border=border
    for cc in (2,3): ws.cell(row=rr,column=cc).alignment=Alignment(horizontal="center")
# live conditional formatting: grey off-rows, green in-rows, colour Net @ fee by sign
ws.conditional_formatting.add(f"A{r0f}:F{rNf}",FormulaRule(formula=[f"$B{r0f}=0"],fill=PatternFill("solid",fgColor="F2F2F2")))
ws.conditional_formatting.add(f"F{r0f}:F{rNf}",FormulaRule(formula=[f"AND($B{r0f}=1,F{r0f}>=0)"],fill=GREEN))
ws.conditional_formatting.add(f"F{r0f}:F{rNf}",FormulaRule(formula=[f"AND($B{r0f}=1,F{r0f}<0)"],fill=ORANGE,font=RED))
for cc in range(1,7): c=ws.cell(row=totf,column=cc);c.fill=NAVY;c.font=white;c.border=border
for cc in range(1,6): ws.cell(row=sh,column=cc).fill=LBLUE;ws.cell(row=sh,column=cc).font=bold
for cc in range(1,6): x=ws.cell(row=sh+1,column=cc);x.fill=BLUE;x.font=white;x.border=border
for j in range(len(FEES)):
    rr=sh+2+j
    for cc in range(1,6): ws.cell(row=rr,column=cc).border=border
    ws.cell(row=rr,column=1).font=bold
ws.cell(row=blr,column=1).font=bold;ws.cell(row=blr,column=4).font=Font(bold=True,color="1F3864")
ws.merge_cells(start_row=note,start_column=1,end_row=note,end_column=6)
c=ws.cell(row=note,column=1);c.font=Font(italic=True,size=9);c.alignment=Alignment(wrap_text=True,vertical="center");ws.row_dimensions[note].height=42
autos(ws,[30,8,8,16,16,14])

# ---------- OPERATING BUDGET (non-conference, full year) ----------
OPS=[
 ("SOCIALS & FORMAL  (trimmed 40%)",[
   ("Welcome / clubs-fair recruitment social","~80 ppl · venue + light food",240),
   ("Fall bonding mixer","team social night",150),
   ("Winter mixer","team social night",150),
   ("Year-end formal / awards banquet","~80 guests · leaner venue/catering (no ticket charge)",2640),
   ("Exec appreciation / transition dinner","~12 execs",180),
 ]),
 ("MERCH & MARKETING",[
   ("Branded banner + tablecloth","one-time, reusable",300),
   ("Recruitment promo","posters, business cards, stickers, clubs fair",200),
   ("Website domain + hosting","annual",100),
   ("Team apparel","bulk order, sold AT COST to members (cash-neutral)",0),
 ]),
 ("ADMIN & AWARDS",[
   ("AMS ratification / club registration","annual",50),
   ("Payment-processing fees","Square ~2.9% on delegate-fee collection",150),
   ("Banquet awards / gavels","superlatives at the year-end formal",200),
   ("Misc / office supplies","",100),
 ]),
]
# which calendar month each operating line lands in (for the live Cash Flow)
op_month={"Welcome / clubs-fair recruitment social":"Sep '26","Fall bonding mixer":"Oct '26",
 "Winter mixer":"Jan '27","Year-end formal / awards banquet":"Mar '27","Exec appreciation / transition dinner":"Apr '27",
 "Branded banner + tablecloth":"Sep '26","Recruitment promo":"Sep '26","Website domain + hosting":"Sep '26",
 "Team apparel":"Sep '26","AMS ratification / club registration":"Sep '26","Payment-processing fees":"Mar '27",
 "Banquet awards / gavels":"Mar '27","Misc / office supplies":"Sep '26"}
ws=wb.create_sheet("Operating Budget")
ws.append(["OPERATING BUDGET 2026-27 – non-conference, full year  ·  club size ~60-100","",""])
ws.append(["EDITABLE control panel: change any yellow Cost cell and the Subtotal, Contingency and Operating Total recalculate — and that total flows into the Scenario Builder's ALL-IN, the gap, and the Cash Flow. CAD. The conference program lives on the other tabs.","",""])
ws.append(["","",""])
ws.append(["Category / Line item","Basis","Cost (CAD)"])
opscats={c for c,_ in OPS}
OPB_MONTH_TERMS={}; ops_net=0; first_item=last_item=None
rw=4
for cat,items in OPS:
    rw+=1; ws.cell(row=rw,column=1,value=cat)
    for it,basis,amt in items:
        rw+=1
        ws.cell(row=rw,column=1,value="   "+it); ws.cell(row=rw,column=2,value=basis)
        cell=ws.cell(row=rw,column=3,value=amt); cell.number_format=MONEY  # numeric, editable
        first_item=first_item or rw; last_item=rw; ops_net+=amt
        OPB_MONTH_TERMS.setdefault(op_month[it],[]).append(f"'Operating Budget'!C{rw}")
rw+=1; sub_row=rw; ws.cell(row=rw,column=1,value="Subtotal (net)"); ws.cell(row=rw,column=3,value=f"=SUM(C{first_item}:C{last_item})").number_format=MONEY
rw+=1; cont_row=rw; ws.cell(row=rw,column=1,value="Contingency 10% (on net operating)"); ws.cell(row=rw,column=3,value=f"=C{sub_row}*0.1").number_format=MONEY
rw+=1; tot_row=rw; ws.cell(row=rw,column=1,value="OPERATING TOTAL"); ws.cell(row=rw,column=3,value=f"=C{sub_row}+C{cont_row}").number_format=MONEY
OPB_TOTAL_CELL=f"'Operating Budget'!C{tot_row}"; OPB_CONTING_CELL=f"'Operating Budget'!C{cont_row}"
conting_ops=round(ops_net*0.10); ops_total=ops_net+conting_ops  # Python copies for any static reference
rw+=2; note_row=rw
ws.cell(row=rw,column=1,value="Note: team apparel is bulk-bought and resold to members at cost, so it nets ~$0 cash. The year-end formal is the single biggest non-conference line and is fully club-funded (no ticket charge). Even so, operating spend stays small next to the travel program.")
# style
ws.merge_cells("A1:C1");ws["A1"].font=Font(bold=True,size=12,color="FFFFFF");ws["A1"].fill=NAVY;ws.row_dimensions[1].height=22
ws.merge_cells("A2:C2");ws["A2"].font=Font(italic=True,size=9);ws["A2"].alignment=Alignment(wrap_text=True);ws.row_dimensions[2].height=40
for cc in range(1,4): x=ws.cell(row=4,column=cc);x.fill=BLUE;x.font=white;x.border=border
for rr in range(5,tot_row+1):
    a=str(ws.cell(row=rr,column=1).value or "")
    if a in opscats:
        for cc in range(1,4): ws.cell(row=rr,column=cc).fill=LBLUE;ws.cell(row=rr,column=cc).font=bold
    elif a=="OPERATING TOTAL":
        for cc in range(1,4): c=ws.cell(row=rr,column=cc);c.fill=NAVY;c.font=white;c.border=border
    elif a.startswith("Subtotal") or a.startswith("Contingency"):
        for cc in range(1,4): ws.cell(row=rr,column=cc).font=Font(italic=True,size=9)
        ws.cell(row=rr,column=3).font=bold
    else:  # editable line-item cost cell
        c=ws.cell(row=rr,column=3);c.fill=YEL;c.border=border
ws.merge_cells(start_row=note_row,start_column=1,end_row=note_row,end_column=3)
c=ws.cell(row=note_row,column=1);c.font=Font(italic=True,size=9);c.alignment=Alignment(wrap_text=True,vertical="center");ws.row_dimensions[note_row].height=42
autos(ws,[46,42,14])

# ---------- FUNDRAISING & NET POSITION ----------
slate_total=sum(parts(k,n)['tot'] for k,n in slate)+SUBS
slate_trips=sum(parts(k,n)['n'] for k,n in slate)
total_cost=slate_total+ops_total
AMS_GRANT=1500          # EDITABLE placeholder — replace with confirmed allotment
MEMBER_FEE=10; MEMBERS=80; member_rev=MEMBER_FEE*MEMBERS   # $10 club joining fee x 80
BAR_NIGHTS=3; bar_low,bar_mid,bar_high=300,500,800
bar_season=bar_mid*BAR_NIGHTS   # $500/night x 3 = $1,500
DINNER=180   # team dinner % night: 30 guests x ~$30/head x 20% back
ws=wb.create_sheet("Fundraising & Net")
ws.append(["FUNDRAISING & NET POSITION 2026-27  ·  all-in club cash picture","","","",""])
ws.append(["Ties it together: TOTAL CLUB COST (conference program + operating) vs FUNDING (AMS grant + delegate fees + bar nights). Aim is to subsidize, not break even — the gap is what the club funds from treasury. CAD.","","","",""])
ws.append(["","","","",""])
ws.append(["A · BAR / PERCENTAGE NIGHTS  (the modelled fundraiser)","","","",""])
ws.append(["Scenario","Net / night","# nights","Season total","How it gets there"])
ws.append(["Conservative",m(bar_low),str(BAR_NIGHTS),m(bar_low*BAR_NIGHTS),"percentage night, ~15% of a light turnout"])
ws.append(["Expected (used below)",m(bar_mid),str(BAR_NIGHTS),m(bar_mid*BAR_NIGHTS),"mix of % night + small cover; ~60-80 out"])
ws.append(["Optimistic",m(bar_high),str(BAR_NIGHTS),m(bar_high*BAR_NIGHTS),"ticketed social, venue deal, strong turnout"])
ws.append(["","","","",""])
ws.append(["Team dinner (20% of spend)",m(DINNER),"1",m(DINNER),"30 guests x ~$30/head x 20% back; range $150-240"])
ws.append(["","","","",""])
ws.append(["B · ALL-IN NET POSITION  (LIVE — driven by the Scenario Builder tab)","","","",""])
def fr_row(label,formula,note):
    ws.append([label,"",formula,"",note]); ws.cell(row=ws.max_row,column=3).number_format=MONEY
fr_row("Conference program (selected slate)",f"={SBREF['slate']}","live — toggle confs on Scenario Builder")
fr_row("Operating budget (socials/training/merch/admin)",f"={SBREF['ops']}","fixed — Operating Budget tab")
fr_row("TOTAL CLUB COST",f"={SBREF['allin']}","what the year costs, all-in")
fr_row("AMS / society grant  (set on Scenario Builder)",f"={SBREF['ams']}","edit on Scenario Builder funding inputs")
fr_row("Club joining fee  ($10 x members)",f"={SBREF['dues']}","one-time membership dues, separate from delegate fees")
fr_row("Bar/percentage nights (season)",f"={SBREF['bar']}","from section A / Scenario Builder")
fr_row("Team dinner fundraiser (20% night)",f"={SBREF['dinner']}","from section A / Scenario Builder")
ws.append(["","","","",""])
ws.append(["Delegate fee","Fee revenue (live trips)","+ grant/dues/bar/dinner","Funding total","Club funds (gap)"])
for f in FEES:
    rev=f"={f}*{SBREF['trips']}"; base=f"={SBREF['base']}"
    funding=f"={f}*{SBREF['trips']}+{SBREF['base']}"; gap=f"={SBREF['allin']}-({f}*{SBREF['trips']}+{SBREF['base']})"
    ws.append([f"${f}/delegate",rev,base,funding,gap])
    for cc in (2,3,4,5): ws.cell(row=ws.max_row,column=cc).number_format=MONEY
ws.append(["","","","",""])
ws.append(["BOTTOM LINE: every number in section B is LIVE from the Scenario Builder — toggle conferences, delegate counts, each conference's per-delegate fee, or any funding input there and this table recalculates. Travel is the cost driver; operating is fixed. Raising one conference's fee by $50 moves the gap by $50 x that conference's delegates. The remaining gap is the planned subsidy, not a shortfall to fix.","","","",""])
ws.merge_cells("A1:E1");ws["A1"].font=Font(bold=True,size=13,color="FFFFFF");ws["A1"].fill=NAVY;ws.row_dimensions[1].height=22
ws.merge_cells("A2:E2");ws["A2"].font=Font(italic=True,size=9);ws["A2"].alignment=Alignment(wrap_text=True);ws.row_dimensions[2].height=28
for rr in range(3,ws.max_row+1):
    a=str(ws.cell(row=rr,column=1).value or "")
    if a.startswith("A ·") or a.startswith("B ·"):
        ws.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=5)
        ws.cell(row=rr,column=1).fill=LBLUE;ws.cell(row=rr,column=1).font=bold
    elif a in ("Scenario","Delegate fee"):
        for cc in range(1,6): x=ws.cell(row=rr,column=cc);x.fill=BLUE;x.font=white;x.border=border
    elif a=="TOTAL CLUB COST":
        for cc in range(1,6): c=ws.cell(row=rr,column=cc);c.fill=NAVY;c.font=white;c.border=border
    elif a.startswith("AMS"):
        for cc in range(1,6): ws.cell(row=rr,column=cc).fill=YEL
        ws.cell(row=rr,column=3).font=bold
    elif a.endswith("/delegate"):
        for cc in range(1,6): ws.cell(row=rr,column=cc).border=border
        ws.cell(row=rr,column=5).font=bold;ws.cell(row=rr,column=5).fill=GREEN
    elif a.startswith("Conference program") or a.startswith("Operating budget") or a.startswith("Bar/percentage") or a.startswith("Team dinner fundraiser") or a.startswith("Club joining"):
        ws.cell(row=rr,column=3).font=bold
    elif a.startswith("BOTTOM LINE"):
        ws.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=5)
        c=ws.cell(row=rr,column=1);c.fill=YEL;c.font=Font(bold=True,color="1F3864");c.alignment=Alignment(wrap_text=True,vertical="center");ws.row_dimensions[rr].height=56
autos(ws,[32,18,18,16,40])

# ---------- REGISTRATION WINDOWS ----------
ws=wb.create_sheet("Registration Windows")
ws.append(["REGISTRATION WINDOWS & DEADLINES – Selected Slate (7 confs)  ·  sourced June 2026","","","",""])
ws.append(["Budget uses each conference's REGULAR-window fee (conservative). Earlier windows save where shown. 2027 portals partly unposted — verify any row marked TBD/est before relying on a date.","","","",""])
ws.append(["","","","",""])
ws.append(["Conference","Window / milestone","Closes","Per-del fee","Status / note"])
REGWIN=[
 ("NCSC – Georgetown",[
   ("Priority","Apr 26, 2026 (CLOSED)","lower","missed — priority wave already shut"),
   ("Regular  (budgeted)","Aug 23, 2026","$110 USD","current open window = the budget rate"),
   ("Final delegate payment","Oct 7, 2026","—","hard balance deadline, all 8 delegates"),
 ]),
 ("CIAC – Cornell",[
   ("Early","Jul 31, 2026","$80 USD","list rate; saves $10/del vs $90 — delegates 11+ free"),
   ("Regular  (budgeted)","after Jul 31, 2026","$90 → $45 USD","50% returning-delegation discount (CONFIRMED) halves all reg — baked into the budget"),
   ("Late / balance due","Oct 9, 2026","$90 USD","list price; billable delegates only (pre-discount)"),
 ]),
 ("RevMUNC – GW   (NEW, founded 2026)",[
   ("Early","Aug 10 – Sep 2, 2026","$60 USD","saves $10/del vs budget; interest form = 10% off delegation"),
   ("Regular  (budgeted)","Sep 3 – Nov 30, 2026","$70 USD","budget rate (conservative)"),
   ("Late","Dec 1 – Feb 12, 2027","$80 USD","avoid — per-del fee steps up"),
   ("Conference","Feb 18-21, 2027","—","DC; falls on reading-week wk2 — verify headcount"),
 ]),
 ("ChoMUN – UChicago",[
   ("Priority","Sep 6, 2026","$95 USD","first-time-school discount; saves $15/del"),
   ("Regular  (budgeted)","Jan 10, 2027","$110 USD","budget rate; +5% if paying by credit card"),
   ("Late","after Jan 10, 2027","higher","avoid — per-del fee steps up"),
   ("Payment deadline","Mar 5, 2027","—","full balance due"),
 ]),
 ("McMUN – McGill   (2027 TBD)",[
   ("Portal opens","~summer 2026","—","watch mcmun.org; 2027 fee table unposted"),
   ("Priority wave","~Sep-Oct 2026 (est)","cheapest","register here when it opens"),
   ("Regular  (budgeted)","TBD","$125 CAD est","estimate — confirm finance@mcmun.org"),
   ("Hotel block cutoff","~Dec 25, 2026","—","Le Centre Sheraton; reconfirm 2027 date"),
   ("Final delegate payment","~late Dec 2026","—","paid with Megabus booking"),
 ]),
 ("ConMUN – Concordia   (2027 TBD)",[
   ("Early","~Sep 2026 (est)","~$110 CAD","2027 portal not yet open"),
   ("Regular  (budgeted)","TBD","$125 CAD","2026 figures used as placeholder"),
   ("Final payment","~Feb 2027","—","paid with Megabus booking"),
 ]),
 ("VICS – UVA   (VICS 31 windows TBD)",[
   ("Early / regular","TBD","$105 USD","windows unposted; $105 used as budget rate"),
   ("Refund cutoff","Jan 19, 2027","—","fees owed in full even if you withdraw after"),
   ("Late registration closes","Mar 1, 2027","—","verify — VICS 31 dates not yet published"),
 ]),
]
confhdrs=set()
for conf,wins in REGWIN:
    ws.append([conf,"","","",""]); confhdrs.add(conf)
    for w,closes,fee,note in wins:
        ws.append(["",w,closes,fee,note])
ws.append(["","","","",""])
ws.append(["EARLY-REGISTRATION SAVINGS (not yet banked in budget)","","","",""])
ws.append(["CIAC early","by Jul 31, 2026","$5/del x 10 billable (post-discount)","= $50 USD","early shaves the already-50%-discounted reg a bit further"])
ws.append(["ChoMUN priority","by Sep 6, 2026","$15/del x 8","= $120 USD","budget assumes the $110 regular rate"])
ws.append(["RevMUNC early","by Sep 2, 2026","$10/del x 8 + 10% off","= $80 USD +","budget assumes the $70 regular rate; interest form adds 10%"])
ws.append(["McMUN / ConMUN priority","when 2027 portals open","unquantified","TBD","priority waves cheapest; amount not yet posted"])
ws.append(["NET: registering in the early/priority windows saves ~$250 USD (~$345 CAD) on the three US confs, before McMUN/ConMUN (CIAC's 50% returning discount is already banked in the slate). The budget deliberately omits the early-window savings so they land as upside, not a number you have to hit.","","","",""])
ws.merge_cells("A1:E1");ws["A1"].font=Font(bold=True,size=12,color="FFFFFF");ws["A1"].fill=NAVY;ws.row_dimensions[1].height=22
ws.merge_cells("A2:E2");ws["A2"].font=Font(italic=True,size=9);ws["A2"].alignment=Alignment(wrap_text=True);ws.row_dimensions[2].height=28
for cc in range(1,6): x=ws.cell(row=4,column=cc);x.fill=BLUE;x.font=white;x.border=border
for rr in range(5,ws.max_row+1):
    a=str(ws.cell(row=rr,column=1).value or "")
    b=str(ws.cell(row=rr,column=2).value or "")
    if a in confhdrs:
        ws.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=5)
        ws.cell(row=rr,column=1).fill=LBLUE;ws.cell(row=rr,column=1).font=bold
    elif a=="EARLY-REGISTRATION SAVINGS (not yet banked in budget)":
        for cc in range(1,6): ws.cell(row=rr,column=cc).fill=GREEN;ws.cell(row=rr,column=cc).font=bold
    elif a.startswith("NET:"):
        ws.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=5)
        c=ws.cell(row=rr,column=1);c.fill=YEL;c.font=Font(bold=True,color="1F3864");c.alignment=Alignment(wrap_text=True,vertical="center");ws.row_dimensions[rr].height=42
    elif b:
        for cc in range(1,6): ws.cell(row=rr,column=cc).border=border
        ws.cell(row=rr,column=5).alignment=Alignment(wrap_text=True)
        if "(budgeted)" in b:
            for cc in range(2,6): ws.cell(row=rr,column=cc).fill=YEL
        ws.cell(row=rr,column=3).font=bold
autos(ws,[26,22,20,14,46])

# ---------- FINANCIAL AID & FEE INCENTIVES ----------
ws=wb.create_sheet("Financial Aid")
ws.append(["FINANCIAL AID & FEE INCENTIVES – per conference + external funding  ·  sourced June 2026","","","","",""])
ws.append(["What each conference offers to offset cost: need-based grants, merit/invoice discounts, and structural discounts (first-time / returning / bulk). Amounts are often unpublished — email the contact to get a figure. Yellow rows = aid you actually apply for.","","","","",""])
ws.append(["","","","","",""])
ws.append(["Conference","Program / incentive","Type","Apply by","Contact","Notes"])
AID=[
 ("HDR","IN THE SELECTED SLATE (7 confs)","","","",""),
 ("APPLY","NCSC – Georgetown","NCSCAid grant","Need-based","Sep 1, 2026","k.lozada@modelun.org","APPLY. Must be registered first. Inaugural fixed pool; dollar amount unpublished — email for a figure."),
 ("APPLY","McMUN – McGill","Merit invoice discount (2 waves)","Merit / delegation","Aug 31 / Nov 16, 2026","finance@mcmun.org","APPLY. Discounts the whole invoice; decisions Sep 15 (priority) / Dec 1 (regular), then 2 wks to pay. Strong target for a Canadian team."),
 ("APPLY","CIAC – Cornell","50% returning-delegation discount (CONFIRMED) + 11+ free","Structural","flag to SG at registration","sg@ciaconline.org","CONFIRMED eligible — last attended CIAC XIII (2022), under the 'CIAC XIV or earlier' cutoff. 50% off ALL reg costs, now BAKED INTO the budget (~$745 CAD saved). 11+ free also in budget. Email SG to confirm it stacks with 11+ free."),
 ("INFO","ChoMUN – UChicago","$50 first-time-school fee + legacy discount","Structural","at registration","dda@chomun.org","$50 first-time fee already in budget. Ask re legacy/hardship pricing. No need-based aid posted."),
 ("INFO","RevMUNC – GW","Interest-form 10% off delegation","Structural","early window ~Sep 2, 2026","secgen.revmunc@gmail.com","Submit interest form for 10% off. New conf — confirm fees/existence; no aid program posted."),
 ("INFO","ConMUN – Concordia","None published (early-bird only)","—","—","sg.conmun@cg-aa.org","No aid; only case-by-case invoice flexibility at Secretariat discretion."),
 ("INFO","VICS – UVA","None published","—","—","sg@iro-vics.org","No aid posted; fees sit behind the registration portal. Email to ask re hardship."),
 ("HDR","NOT IN CURRENT SLATE (reference)","","","",""),
 ("APPLY","UNCMUNC – UNC","Need-based aid + CIDP 50%","Need-based + structural","~Jun 22, 2026","uncmunc@gmail.com","Applies ONLY IF swapped into the slate. Need-based open to all; CIDP 50% is first-time-Southern/HBCU/MSI only (we likely do not qualify)."),
 ("INFO","UPMUNC – Penn","None (timing tiers only)","—","—","secretariat","No aid; priority-window pricing only."),
 ("INFO","BUCKMUN – Ohio State","Unverifiable","—","—","buckmun@ccwaosu.org","No web presence; OSU's real conf is OSUMUN (no aid). Confirm this conference exists before relying on it."),
]
aidkind={}
for row in AID:
    ws.append(list(row[1:])); aidkind[ws.max_row]=row[0]
ws.append(["","","","","",""])
ws.append(["EXTERNAL / INSTITUTIONAL FUNDING (offsets any conference; stackable)","","","","",""])
ws.append(["AMS Clubs Experience Grant","up to $4,000 / period","Internal grant","rolling (Summer/Fall/Winter)","myams.org/clubs","Largest internal pot. Pitch a trip as a club initiative; needs an AMS bank account."])
ws.append(["Principal's Student Initiatives Fund","conferences / competitions","Internal grant","rolling, ~2-3 day decision","queensu.ca/principal","Cannot tie to course credit; frame as the delegation."])
ws.append(["FAS initiative funds (Anderson / Bartlett)","program-gated","Internal grant","Oct / Feb / May","queensu.ca/artsci","Eligibility depends on members' majors (Classics/English/Phil/History etc.)."])
ws.append(["No dedicated Canadian MUN travel grant","—","note","—","—","UNA-Canada/CANIMUN and ACUNS do NOT fund delegation travel to US confs. The money is internal + direct department asks."])
ws.append(["","","","","",""])
ws.append(["NET: McMUN (merit invoice discount) and NCSC (NCSCAid) are the two grants worth applying to — both deadline-driven Aug-Sep 2026. CIAC's 50% returning-delegation discount is CONFIRMED and already baked into the slate (~$745 CAD). UNCMUNC need-based opens a third grant only if it is swapped in (apply by ~Jun 22). Everything else is register-early structural savings tracked on Registration Windows.","","","","",""])
ws.merge_cells("A1:F1");ws["A1"].font=Font(bold=True,size=12,color="FFFFFF");ws["A1"].fill=NAVY;ws.row_dimensions[1].height=22
ws.merge_cells("A2:F2");ws["A2"].font=Font(italic=True,size=9);ws["A2"].alignment=Alignment(wrap_text=True);ws.row_dimensions[2].height=30
for cc in range(1,7): x=ws.cell(row=4,column=cc);x.fill=BLUE;x.font=white;x.border=border
for rr in range(5,ws.max_row+1):
    a=str(ws.cell(row=rr,column=1).value or "")
    kind=aidkind.get(rr)
    if kind=="HDR":
        ws.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=6)
        ws.cell(row=rr,column=1).fill=LBLUE;ws.cell(row=rr,column=1).font=bold
    elif a=="EXTERNAL / INSTITUTIONAL FUNDING (offsets any conference; stackable)":
        ws.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=6)
        ws.cell(row=rr,column=1).fill=GREEN;ws.cell(row=rr,column=1).font=bold
    elif a.startswith("NET:"):
        ws.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=6)
        c=ws.cell(row=rr,column=1);c.fill=YEL;c.font=Font(bold=True,color="1F3864");c.alignment=Alignment(wrap_text=True,vertical="center");ws.row_dimensions[rr].height=46
    elif a:
        for cc in range(1,7): ws.cell(row=rr,column=cc).border=border
        ws.cell(row=rr,column=6).alignment=Alignment(wrap_text=True)
        ws.cell(row=rr,column=1).font=bold;ws.cell(row=rr,column=4).font=bold
        if kind=="APPLY":
            for cc in range(1,7): ws.cell(row=rr,column=cc).fill=YEL
autos(ws,[24,30,18,20,26,52])

# ---------- CASH FLOW (month-by-month timing, LIVE from the Scenario Builder) ----------
# Each conference's deposit/balance/fee revenue is a formula off the Scenario Builder, gated by its
# Go toggle — so toggling confs, delegate counts, the fee, FX, or funding inputs re-times the whole
# schedule. Month ASSIGNMENT (which month a payment lands) is a fixed planning assumption.
# Year-end cumulative == -(gap): inflows - outflows reduces to FUNDING - ALL-IN, an internal check.
MONTHS=["Jul–Aug '26","Sep '26","Oct '26","Nov '26","Dec '26","Jan '27","Feb '27","Mar '27","Apr '27","Post-season"]
outT={mn:[] for mn in MONTHS}; innT={mn:[] for mn in MONTHS}; drv={mn:[] for mn in MONTHS}
def OUT(mn,term,why): outT[mn].append(term); drv[mn].append("− "+why)
def IN(mn,term,why): innT[mn].append(term); drv[mn].append("+ "+why)
FX_ABS=f"'Scenario Builder'!$D$3"  # per-conf delegate fee now lives in col F of each SB row
# (deposit month, settlement month) by conference calendar position — covers all 10 menu confs
cf_sched={"UPMUNC":("Jul–Aug '26","Oct '26"),"CIAC":("Jul–Aug '26","Oct '26"),"NCSC":("Sep '26","Nov '26"),
          "UNCMUNC":("Sep '26","Nov '26"),"MCMUN":("Sep '26","Dec '26"),"BUCKMUN":("Sep '26","Feb '27"),
          "REVMUNC":("Sep '26","Feb '27"),"CONMUN":("Sep '26","Feb '27"),"CHOMUN":("Sep '26","Mar '27"),
          "VICS":("Jan '27","Mar '27")}
for k in SB_ORDER:
    c=CONF[k]; sbr=SB_ROW[k]; dmn,smn=cf_sched[k]; short=c['full'].split(" – ")[0].split(" ")[0]
    go=f"'Scenario Builder'!B{sbr}"; dele=f"'Scenario Builder'!C{sbr}"; ec=f"'Scenario Builder'!E{sbr}"
    dep=f"{c['deleg']}*{FX_ABS}*{go}" if c['cur']=="USD" else f"{c['deleg']}*{go}"  # delegation deposit (CAD), gated by Go
    OUT(dmn,dep,f"{short} deposit"); OUT(smn,f"({ec}-({dep}))",f"{short} balance")
    IN(smn,f"'Scenario Builder'!F{sbr}*{dele}*{go}",f"{short} delegate fees")
# transport bonds (refundable) + subscriptions, then funding inputs pulled live off the Scenario Builder
OUT("Oct '26","1500","Communauto bonds (3×$500, refundable)"); IN("Post-season","1500","Communauto bonds returned")
OUT("Oct '26",str(int(SUBS)),"Communauto Value subscriptions")
IN("Sep '26",f"'Scenario Builder'!E{_SB_MEM}","member dues"); IN("Oct '26",f"'Scenario Builder'!E{_SB_AMS}","AMS / society grant")
BAR3=f"'Scenario Builder'!E{_SB_BAR}/3"
IN("Nov '26",BAR3,"bar / % night (fall)"); IN("Feb '27",BAR3,"bar / % night (winter)")
IN("Mar '27",BAR3,"bar / % night (spring)"); IN("Mar '27",f"'Scenario Builder'!E{_SB_DIN}","team dinner (20% night)")
# operating outflows: each month references its live Operating Budget line cells (+ contingency in Apr)
for mn,terms in OPB_MONTH_TERMS.items(): OUT(mn,"+".join(terms),"operating (socials/merch/admin)")
OUT("Apr '27",OPB_CONTING_CELL,"operating contingency")
ws=wb.create_sheet("Cash Flow")
ws.append(["CASH FLOW 2026-27 — month-by-month money out vs money in  ·  LIVE from the Scenario Builder","","","","",""])
ws.append(["Shows WHEN cash moves, not just the annual total. The cumulative column is the point: payments land before the grant and most delegate fees arrive, so the club needs a treasury buffer to bridge the low point. Toggle confs / counts / fee / funding on the Scenario Builder and this re-times itself. Month assignment is a fixed planning assumption. CAD.","","","","",""])
ws.append(["","","","","",""])
ws.append(["Month","Cash out","Cash in","Net month","Cumulative","Key drivers (what lands this month)"])
r0c=5
for i,mn in enumerate(MONTHS):
    rr=r0c+i
    of="="+("+".join(outT[mn]) if outT[mn] else "0"); inf="="+("+".join(innT[mn]) if innT[mn] else "0")
    ws.cell(row=rr,column=1,value=mn)
    ws.cell(row=rr,column=2,value=of).number_format=MONEY
    ws.cell(row=rr,column=3,value=inf).number_format=MONEY
    ws.cell(row=rr,column=4,value=f"=C{rr}-B{rr}").number_format=MONEY
    ws.cell(row=rr,column=5,value=(f"=D{rr}" if i==0 else f"=E{rr-1}+D{rr}")).number_format=MONEY
    ws.cell(row=rr,column=6,value="  ".join(dict.fromkeys(drv[mn]))[:160])
rNc=r0c+len(MONTHS)-1
lowr=rNc+2; endr=lowr+1
ws.cell(row=lowr,column=1,value="WORKING-CAPITAL LOW POINT (most negative cumulative) →")
ws.cell(row=lowr,column=5,value=f"=MIN(E{r0c}:E{rNc})").number_format=MONEY
ws.cell(row=endr,column=1,value="Year-end position (= the planned subsidy / Fundraising gap) →")
ws.cell(row=endr,column=5,value=f"=E{rNc}").number_format=MONEY
noterow=endr+2
ws.cell(row=noterow,column=1,value="The low point is the treasury buffer (or AMS advance / bridge) the club must hold before reimbursements and spring fees land. Timing model: delegation deposits at registration; each conference's balance + delegate fees at settlement (McMUN settles Dec on the hotel-block cutoff; VICS deposit in Jan at the refund cutoff). AMS grant assumed Oct; dues at the fall clubs fair; bar nights one per term; formal in March. Edit those months to your actual portal/AMS dates in the build script.")
ws.merge_cells("A1:F1");ws["A1"].font=Font(bold=True,size=13,color="FFFFFF");ws["A1"].fill=NAVY;ws.row_dimensions[1].height=22
ws.merge_cells("A2:F2");ws["A2"].font=Font(italic=True,size=9);ws["A2"].alignment=Alignment(wrap_text=True);ws.row_dimensions[2].height=48
for cc in range(1,7): x=ws.cell(row=4,column=cc);x.fill=BLUE;x.font=white;x.border=border;x.alignment=Alignment(wrap_text=True,vertical="center")
for i,mn in enumerate(MONTHS):
    rr=r0c+i
    for cc in range(1,7): ws.cell(row=rr,column=cc).border=border
    ws.cell(row=rr,column=6).alignment=Alignment(wrap_text=True);ws.row_dimensions[rr].height=30
    ws.cell(row=rr,column=4).font=bold
    if mn=="Post-season":
        for cc in range(1,7): ws.cell(row=rr,column=cc).fill=GREEN
# live: highlight whichever cumulative cell is the minimum
ws.conditional_formatting.add(f"E{r0c}:E{rNc}",FormulaRule(formula=[f"E{r0c}=MIN($E${r0c}:$E${rNc})"],fill=ORANGE,font=RED))
for rr,fill in ((lowr,YEL),(endr,LBLUE)):
    for cc in range(1,6): c=ws.cell(row=rr,column=cc);c.fill=fill;c.border=border
    ws.cell(row=rr,column=1).font=Font(bold=True,color="1F3864");ws.cell(row=rr,column=5).font=Font(bold=True,color="1F3864")
ws.merge_cells(start_row=noterow,start_column=1,end_row=noterow,end_column=6)
c=ws.cell(row=noterow,column=1);c.font=Font(italic=True,size=9);c.alignment=Alignment(wrap_text=True,vertical="center");ws.row_dimensions[noterow].height=56
autos(ws,[14,13,13,13,13,52])

# ---------- SCENARIO BUILDER (live toggle tab) ----------
# Chronological menu of every conference + preset = current 7-conf slate.
# (key, preset Go 1/0, preset delegates, Montreal-fixed?)
SB=[("UPMUNC",0,8,False),("CIAC",1,12,False),("NCSC",1,8,False),("UNCMUNC",0,8,False),
    ("MCMUN",1,20,True),("BUCKMUN",0,8,False),("REVMUNC",1,8,False),("CONMUN",1,20,True),
    ("CHOMUN",1,8,False),("VICS",1,12,False)]
assert [k for k,*_ in SB]==SB_ORDER,"SB menu order drifted from SB_ORDER (Selected Slate references it)"
OPS_TOTAL=4906; BASE_FUNDING=3980  # from Operating Budget + Fundraising tabs
ws=wb.create_sheet("Scenario Builder",1)
ws.append(["SCENARIO BUILDER — toggle conferences, the budget updates live","","","","","","","","","","",""])
ws.append(["Set Go? to 1 (attend) or 0 (skip). Set Delegates (US: 8 or 12 · Montreal: 16 or 20). Set Fee /del to what each delegate pays for THAT conference. Yellow cells are yours to edit; everything below recalculates automatically.","","","","","","","","","","",""])
ws.append(["FX rate  USD → CAD  (edit — scales only the US-conf costs) →","","",1.39,"","","","","","","","","",""])
ws.append(["Conference","Go?","Delegates","Cost @ size (CAD)","In-slate cost","Fee /del","usd@8","cad@8","usd@12","cad@12","usd@16","cad@16","usd@20","cad@20"])
FXC="$D$3"  # editable FX cell; conference costs scale their USD part by this, CAD part is fixed
HR=4; r0=HR+1
for i,(k,go,dele,mtl) in enumerate(SB):
    r=r0+i
    splits=[split_at(k,8),split_at(k,12),split_at(k,16),split_at(k,20)]  # refs at every selectable size
    ws.cell(row=r,column=1,value=CONF[k]['full']+("  [16 or 20]" if mtl else ""))
    ws.cell(row=r,column=2,value=go)
    ws.cell(row=r,column=3,value=dele)
    usd=f"IF(C{r}=8,G{r},IF(C{r}=12,I{r},IF(C{r}=16,K{r},M{r})))"
    cad=f"IF(C{r}=8,H{r},IF(C{r}=12,J{r},IF(C{r}=16,L{r},N{r})))"
    ws.cell(row=r,column=4,value=f"=({usd}*{FXC}+{cad})*{CONTING}").number_format=MONEY
    ws.cell(row=r,column=5,value=f"=B{r}*D{r}").number_format=MONEY
    ws.cell(row=r,column=6,value=300).number_format=MONEY  # per-conference delegate fee (editable)
    for col,(u,cd) in zip((7,9,11,13),splits):   # usd parts in G/I/K/M, cad parts in H/J/L/N
        for cc_col,val in ((col,u),(col+1,cd)):
            cc=ws.cell(row=r,column=cc_col,value=round(val)); cc.number_format=MONEY; cc.font=Font(size=9,color="808080")
rN=r0+len(SB)-1
subs=rN+1
assert (r0,subs)==(_SB_R0,_SB_SUBS),"SB conf-row block drifted; update _SB_R0/_SB_SUBS (Selected Slate references them)"
ws.cell(row=subs,column=1,value="+ Communauto Value subscriptions (3 accts, if driving)")
ws.cell(row=subs,column=5,value=180).number_format=MONEY
blank=subs+1
SLATE=blank+1; OPSR=SLATE+1; ALLIN=OPSR+1
ws.cell(row=SLATE,column=1,value="SLATE TOTAL")
ws.cell(row=SLATE,column=2,value=f"=SUMPRODUCT(B{r0}:B{rN},C{r0}:C{rN})")  # delegate-trips
ws.cell(row=SLATE,column=3,value="trips")
ws.cell(row=SLATE,column=4,value=f'=IF(B{SLATE}=0,"",E{SLATE}/B{SLATE})').number_format=MONEY  # per-trip
ws.cell(row=SLATE,column=5,value=f"=SUM(E{r0}:E{subs})").number_format=MONEY
ws.cell(row=OPSR,column=1,value="Operating budget (live — edit on Operating Budget tab)")
ws.cell(row=OPSR,column=5,value=f"={OPB_TOTAL_CELL}").number_format=MONEY
ws.cell(row=ALLIN,column=1,value="ALL-IN TOTAL")
ws.cell(row=ALLIN,column=5,value=f"=E{SLATE}+E{OPSR}").number_format=MONEY
# funding inputs (all editable) — this control panel drives the Fundraising tab
FHDR=ALLIN+2; INP_AMS=FHDR+1; INP_MEM=FHDR+2; INP_BAR=FHDR+3; INP_DIN=FHDR+4; BASER=FHDR+5
FEER=BASER+2; FUNDR=FEER+1; GAPR=FEER+2
ws.cell(row=FHDR,column=1,value="FUNDING INPUTS  — edit any yellow cell")
ws.cell(row=INP_AMS,column=1,value="AMS / society grant")
ws.cell(row=INP_AMS,column=5,value=1500).number_format=MONEY
ws.cell(row=INP_MEM,column=1,value="Club members  (x $10 joining fee)")
ws.cell(row=INP_MEM,column=4,value=80)
ws.cell(row=INP_MEM,column=5,value=f"=D{INP_MEM}*10").number_format=MONEY
ws.cell(row=INP_BAR,column=1,value="Bar / percentage nights (season total)")
ws.cell(row=INP_BAR,column=5,value=1500).number_format=MONEY
ws.cell(row=INP_DIN,column=1,value="Team dinner (20% night)")
ws.cell(row=INP_DIN,column=5,value=180).number_format=MONEY
ws.cell(row=BASER,column=1,value="Base funding (auto)")
ws.cell(row=BASER,column=5,value=f"=E{INP_AMS}+E{INP_MEM}+E{INP_BAR}+E{INP_DIN}").number_format=MONEY
ws.cell(row=FEER,column=1,value="Delegate fees (set per conf in 'Fee /del' col) →")
ws.cell(row=FEER,column=3,value="blended /del →")
ws.cell(row=FEER,column=4,value=f'=IF(B{SLATE}=0,"",E{FEER}/B{SLATE})').number_format=MONEY
ws.cell(row=FEER,column=5,value=f"=SUMPRODUCT(B{r0}:B{rN},C{r0}:C{rN},F{r0}:F{rN})").number_format=MONEY
ws.cell(row=FUNDR,column=1,value="FUNDING TOTAL")
ws.cell(row=FUNDR,column=5,value=f"=E{BASER}+E{FEER}").number_format=MONEY
ws.cell(row=GAPR,column=1,value="CLUB FUNDS (GAP) — the planned subsidy")
ws.cell(row=GAPR,column=5,value=f"=E{ALLIN}-E{FUNDR}").number_format=MONEY
BEVENR=GAPR+1
ws.cell(row=BEVENR,column=1,value="Break-even delegate fee  (fee that drives the gap to $0)")
ws.cell(row=BEVENR,column=5,value=f'=IF(B{SLATE}=0,"",(E{ALLIN}-E{BASER})/B{SLATE})').number_format=MONEY
# safety: the Fundraising tab references the addresses below; fail loudly if the layout drifts
assert (SLATE,OPSR,ALLIN,INP_AMS,INP_MEM,INP_BAR,INP_DIN,BASER,FEER,FUNDR,GAPR,BEVENR)==(
    _SB_SLATE,_SB_OPS,_SB_ALLIN,_SB_AMS,_SB_MEM,_SB_BAR,_SB_DIN,_SB_BASE,_SB_FEE,_SB_FUND,_SB_GAP,_SB_BEVEN),"Scenario Builder layout drifted; update _SB_* constants"
# data validation
dv_go=DataValidation(type="list",formula1='"0,1"',allow_blank=False); ws.add_data_validation(dv_go)
dv_us=DataValidation(type="list",formula1='"8,12"',allow_blank=False); ws.add_data_validation(dv_us)
dv_mtl=DataValidation(type="list",formula1='"16,20"',allow_blank=False); ws.add_data_validation(dv_mtl)
for i,(k,go,dele,mtl) in enumerate(SB):
    r=r0+i; dv_go.add(ws.cell(row=r,column=2)); (dv_mtl if mtl else dv_us).add(ws.cell(row=r,column=3))
# conditional formatting: included rows green, skipped grey
ws.conditional_formatting.add(f"A{r0}:E{rN}",FormulaRule(formula=[f"$B{r0}=1"],fill=GREEN))
ws.conditional_formatting.add(f"A{r0}:E{rN}",FormulaRule(formula=[f"$B{r0}=0"],fill=PatternFill("solid",fgColor="F2F2F2")))
# styling
ws.merge_cells("A1:N1");ws["A1"].font=Font(bold=True,size=14,color="FFFFFF");ws["A1"].fill=NAVY;ws.row_dimensions[1].height=24
ws.merge_cells("A2:N2");ws["A2"].font=Font(italic=True,size=9);ws["A2"].alignment=Alignment(wrap_text=True);ws.row_dimensions[2].height=28
ws.merge_cells("A3:C3");ws.cell(row=3,column=1).font=Font(bold=True,size=10)
fxcell=ws.cell(row=3,column=4);fxcell.number_format="0.00";fxcell.fill=YEL;fxcell.font=Font(bold=True);fxcell.border=border;fxcell.alignment=Alignment(horizontal="center")
for cc in range(1,15): x=ws.cell(row=HR,column=cc);x.fill=BLUE;x.font=white;x.border=border;x.alignment=Alignment(wrap_text=True,vertical="center")
for i in range(len(SB)):
    r=r0+i
    for cc in range(1,7): ws.cell(row=r,column=cc).border=border
    for cc in (2,3,6): ws.cell(row=r,column=cc).fill=YEL  # editable cells (Go, Delegates, Fee /del)
for rr,fill in ((SLATE,LBLUE),(ALLIN,NAVY),(FUNDR,LBLUE)):
    for cc in range(1,6): c=ws.cell(row=rr,column=cc);c.fill=fill;c.font=(white if fill==NAVY else bold);c.border=border
ws.merge_cells(start_row=FHDR,start_column=1,end_row=FHDR,end_column=5)
ws.cell(row=FHDR,column=1).fill=LBLUE;ws.cell(row=FHDR,column=1).font=bold
ws.cell(row=INP_AMS,column=5).fill=YEL; ws.cell(row=INP_MEM,column=4).fill=YEL
ws.cell(row=INP_BAR,column=5).fill=YEL; ws.cell(row=INP_DIN,column=5).fill=YEL
ws.cell(row=BASER,column=5).font=bold
ws.cell(row=FEER,column=4).fill=YEL  # fee input editable
for cc in range(1,6): c=ws.cell(row=GAPR,column=cc);c.fill=YEL;c.font=Font(bold=True,color="C00000");c.border=border
for cc in range(1,6): c=ws.cell(row=BEVENR,column=cc);c.fill=GREEN;c.font=bold;c.border=border
# risk note: FX cushion + no-show / sunk-cost exposure
NOTE1=BEVENR+2
ws.cell(row=NOTE1,column=1,value="FX cushion: everything US is priced at 1.39. Bump the FX cell to 1.45 and the slate cost rises ~"+m(sum(parts_split(k,n)[0] for k,n in slate)*0.06*CONTING)+" — that is the model's single biggest external risk, now a live lever above.")
ws.cell(row=NOTE1+1,column=1,value="No-show / sunk-cost: the guaranteed-delegate assumption is load-bearing. Most delegation deposits and hotel blocks are NON-refundable once paid, so an unfilled seat is money already spent. Register conservatively and confirm headcount before each deposit deadline (see Payment Timeline).")
for rr in (NOTE1,NOTE1+1):
    ws.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=14)
    c=ws.cell(row=rr,column=1);c.font=Font(italic=True,size=9,color="843C0C");c.fill=ORANGE;c.alignment=Alignment(wrap_text=True,vertical="center");ws.row_dimensions[rr].height=30
autos(ws,[34,7,11,17,15,10,9,9,9,9,9,9,9,9])

path=os.path.expanduser("~/Downloads/Queens_MUN_Budget_2026-27.xlsx")
wb.save(path)
print("SAVED",path,os.path.getsize(path)); print("TABS:",wb.sheetnames)
print(f"8-scn program total: C${g8+SUBS:,.0f} (US${(g8+SUBS)/USD_CAD:,.0f})")
print(f"12-scn program total: C${g12+SUBS:,.0f} (US${(g12+SUBS)/USD_CAD:,.0f})")
st=sum(parts(k,n)['tot'] for k,n in slate)+SUBS
print(f"Selected slate: C${st:,.0f} (US${st/USD_CAD:,.0f})")
print("McMUN@20:",m(parts('MCMUN',8)['tot']),"ConMUN@20:",m(parts('CONMUN',8)['tot']))
