"""Fill the QIAA Budget Proposal template with the 2026-27 Queen's MUN travel slate.

Reads the QIAA template (two sheets: 'budget request' + 'portfolio summary'),
preserves the QIAA instructions block + headers, and writes the MUN program as
flat expense rows broken into registration / hotel / transport per conference,
plus operating lines. Numbers are pulled from the authoritative workbook
(Queens_MUN_Budget_2026-27.xlsx) selected slate (7 confs): CIAC 12, NCSC 8,
McMUN 20, RevMUNC 8, ConMUN 20, ChoMUN 8, VICS 12.

Framing (per Jack): QIAA approves the FULL ~$50,094 program; delegate fees +
dues + fundraising cover ~$30,380; the remaining ~$19,714 is the planned
operating loss the club asks QIAA to backstop. All figures CAD.

Output: ~/Downloads/QIAA 2026-2027 - MUN Budget Proposal (Queen's MUN - FILLED).xlsx
"""
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

TEMPLATE = "/Users/jackguillemette/Downloads/QIAA 2026–2027 - MUN Budget Proposal.xlsx"
OUT = os.path.expanduser(
    "~/Downloads/QIAA 2026-2027 - MUN Budget Proposal (Queen's MUN - FILLED).xlsx"
)

# ---- plain, neutral palette (looks like a normally filled-in form) ----
NAVY = "404040"      # header text/fill: dark grey, not branded navy
RED = "404040"
GREYBG = "F2F2F2"
ZEBRA = "F8F8F8"
HDR_FILL = "D9D9D9"  # light grey header
TOTAL_FILL = "FFF2CC"  # soft highlight on the total row

# ---- expense rows -------------------------------------------------------
# (Expense Item, Category, Cost CAD, Explanation, Necessity 1-5, Consequence, Source)
# Conference rows are grouped reg / hotel / transport. Necessity is per-conference
# (you cannot attend without all three lines).

CONF_ROWS = []

def conf(name, nec, lines):
    # lines: list of (suffix, cost, explanation, consequence)
    for suffix, cost, expl, cons in lines:
        CONF_ROWS.append((
            f"{name}: {suffix}", "Conference travel", cost, expl, nec, cons, SRC[name],
        ))

SRC = {
    "CIAC (Cornell)": "ciaconline.org/register; Cayuga Blu Hotel block; Communauto Value + Long Distance. Sourced Jun 2026.",
    "NCSC (Georgetown)": "ncsc.modelun.org/registration; Washington Hilton group rate; Communauto. Regular window (closes Aug 23, 2026).",
    "McMUN (McGill)": "mcmun.org (2027 fee table TBD, $125/del CAD est); Le Centre Sheraton quad $234 CAD; Megabus Kingston-Montreal RT $110.",
    "RevMUNC (GW)": "gwuias.org/revmunc (new conf 2026); Foggy Bottom hotel est (no block posted); Communauto. Regular window.",
    "ConMUN (Concordia)": "conmun.org/en/registration (2026 figures, 2027 TBD); DoubleTree Montreal $229 CAD; Megabus RT $110.",
    "ChoMUN (UChicago)": "chomun.org/registration; Palmer House Hilton $201/quad; Communauto. Regular window (+5% if paid by card).",
    "VICS (UVA)": "iro-vics.org (VICS 31 windows TBD, $105/del USD used); Charlottesville hotels $169-222; Communauto.",
}

conf("CIAC (Cornell)", 4, [
    ("Registration", 683,
     "Cornell, late October, 12 delegates. We get the 50% returning-school discount (we last went in 2022) and anyone past 11 delegates is free, so we only pay for 10. Cheapest registration on the slate.",
     "The returning discount only applies if we actually go, so we'd lose it."),
    ("Hotel", 1877,
     "3 rooms, 3 nights at the Cayuga Blu. Rate includes breakfast.",
     "Nowhere to put the delegates, so the trip is off."),
    ("Transport", 1118,
     "3 Communauto cars, about 3.5 hours each way. Gas and tax included.",
     "No way to get the team to Ithaca."),
])

conf("NCSC (Georgetown)", 5, [
    ("Registration", 1358,
     "Hosted by Georgetown and one of the strongest conferences we attend, 8 delegates. Regular registration is $90 for the school plus $110 per delegate.",
     "Skipping a top-tier conference hurts us competitively and it's a real draw when recruiting."),
    ("Hotel", 3074,
     "2 rooms, 3 nights at the Washington Hilton block. Most expensive hotel on the slate, but it's downtown DC.",
     "No rooms means no delegation in Washington."),
    ("Transport", 1273,
     "2 cars, roughly 8 hours down.",
     "Delegates can't reach DC."),
])

conf("McMUN (McGill)", 5, [
    ("Registration", 2080,
     "Our flagship trip and the one everyone wants to be on. McGill caps schools at 20; we're planning 16 this year. The 2027 fee table isn't out yet, so $125/delegate is based on last year.",
     "Pulling out of McMUN would be the biggest setback to the team's year."),
    ("Hotel", 2808,
     "4 quad rooms, 3 nights at Le Centre Sheraton.",
     "No housing for the delegation."),
    ("Transport", 1760,
     "16 Megabus round trips, Kingston to Montreal, about 3.5 hours.",
     "Team can't get to Montreal."),
])

conf("RevMUNC (GW)", 3, [
    ("Registration", 868,
     "New conference (first ran in 2026), at George Washington, 8 delegates. It lands on the second week of reading week, so we need to confirm headcount. Budgeted at the regular-window fee.",
     "We'd skip a cheap DC weekend. Not a big loss if our numbers are tight that month."),
    ("Hotel", 1932,
     "2 rooms, 3 nights, Foggy Bottom area. No block posted yet so this is an estimate.",
     "No rooms, no trip."),
    ("Transport", 1273,
     "Same DC drive as NCSC, 2 cars.",
     "No way to get there."),
])

conf("ConMUN (Concordia)", 4, [
    ("Registration", 2100,
     "Concordia's conference in Montreal, 16 delegates. The 2027 dates and fees aren't posted, so these are last year's figures.",
     "Losing an easy, high-capacity Montreal trip."),
    ("Hotel", 2748,
     "4 quad rooms, 3 nights at the DoubleTree.",
     "No housing for the delegation."),
    ("Transport", 1760,
     "16 Megabus round trips.",
     "Can't get the team to Montreal."),
])

conf("ChoMUN (UChicago)", 5, [
    ("Registration", 1302,
     "UChicago, one of the best-run conferences out there, 8 delegates. $50 school fee plus $110 per delegate. Note they tack on 5% if we pay by card.",
     "Another top conference our stronger delegates target."),
    ("Hotel", 1688,
     "2 rooms, 3 nights at the Palmer House Hilton.",
     "No rooms in Chicago."),
    ("Transport", 1567,
     "2 cars, our longest drive at around 10 hours each way.",
     "No way to reach Chicago."),
])

conf("VICS (UVA)", 3, [
    ("Registration", 1883,
     "UVA's conference, 12 delegates. VICS 31 hasn't posted its 2027 windows, so $105/delegate is our working number.",
     "Drops a southern trip. Manageable if the date ends up clashing with ChoMUN."),
    ("Hotel", 2381,
     "3 rooms, 3 nights. Charlottesville hotels run $169-222; used $189.",
     "No housing for the trip."),
    ("Transport", 2237,
     "3 cars, 11 to 13 hours each way.",
     "Too far to reach without the cars."),
])

# travel-program tail
TRAVEL_TAIL = [
    ("Communauto Value subscriptions (3 accounts)", "Conference travel", 180,
     "3 Value accounts at $60/year; the $5/mo plan unlocks Long Distance pricing, which makes US drives feasible. Shared across all driving conferences.", 4,
     "Lose Long Distance rates; per-trip car cost rises sharply or US conferences become undrivable.",
     "Communauto Value plan + Long Distance, communauto.com. 3 x $500 refundable bonds are tied up but are not a cost."),
    ("Travel program contingency (10%)", "Contingency", 3777,
     "10% buffer on the conference travel subtotal for FX moves, hotel-rate changes, and registration-fee revisions on the conferences with unposted 2027 portals.", 3,
     "No cushion against the USD/CAD rate (priced at 1.40) or the several estimated/TBD fees; an overrun would come straight off the club treasury.",
     "Standard 10% contingency on subtotal; FX 1.40 is the model's single biggest external risk."),
]

OPERATING_ROWS = [
    ("Fall initiation social", "Socials", 200,
     "Start-of-year initiation social, about 80 people, refreshments. Top of the funnel for new members.", 4,
     "Weaker recruitment and a smaller delegate pool, so fewer paying delegates to offset travel.",
     "~80 people, refreshments."),
    ("Fall social (midterm watch party)", "Socials", 100,
     "Team social night around midterms.", 2,
     "Team is a bit less close. Not essential.",
     "Social night estimate."),
    ("Winter social", "Socials", 200,
     "Team social night in the new year.", 2,
     "Nice to have, not essential.",
     "Social night estimate."),
    ("MUNDIES (year-end awards night)", "Socials", 300,
     "Year-end awards night, leaner venue and catering, no ticket charge.", 3,
     "No capstone event. Could shrink it or go ticketed if the money isn't there.",
     "Leaner venue/catering, club-funded."),
    ("Recruitment (Queen's in the Park)", "Marketing", 75,
     "Table and presence at the Queen's in the Park recruitment event.", 3,
     "Miss a key recruiting touchpoint at the start of the year.",
     "Event table/materials."),
    ("Recruitment promo", "Marketing", 75,
     "Posters, business cards, stickers.", 3,
     "Less reach when recruiting.",
     "Print/promo estimate."),
    ("App / software costs", "Software", 300,
     "Annual app and software costs for the team (site, tools).", 3,
     "Lose the tools the team runs on.",
     "Annual, software/hosting."),
    ("Team apparel", "Merch", 0,
     "Bulk order sold to members at cost, so it nets out to zero. Listed just so it's on the record.", 1,
     "None. Members pay for their own.",
     "Resold at cost (nets ~$0)."),
    ("Banquet awards / gavels", "Awards", 200,
     "Superlatives and gavels handed out at the year-end night.", 2,
     "No physical awards. Low impact.",
     "Awards/gavels estimate."),
    ("Misc / office supplies", "Admin", 100,
     "Printing, placards, odds and ends over the year.", 2,
     "Minor friction.",
     "General supplies."),
    ("Operating contingency (20%)", "Contingency", 310,
     "20% buffer on the operating lines.", 3,
     "No cushion if a number comes in high.",
     "20% on operating subtotal."),
]

# funding-context rows (NOT expenses; documented below the table for transparency)
FUNDING_CONTEXT = [
    ("Projected delegate fees (80 trips, blended ~$308)", -24600,
     "Delegates pay a fee per conference, set so the cheaper trips help carry the expensive ones. 80 delegate-trips brings in about $24,600."),
    ("Club joining dues (80 members x $10)", -800,
     "One-time $10 to join, separate from the per-conference fees."),
    ("Fundraising (bar/percentage nights + team dinners)", -2250,
     "3 bar/percentage nights (~$2,000) plus two team-dinner fundraisers (~$250)."),
    ("AMS / society grant (placeholder)", -1500,
     "Placeholder until the AMS/society number is confirmed."),
]


QIAA_INTRO = (
    "QIAA Budget Proposal Submission\n\n"
    "As part of the budgeting process for the 2026–2027 year, each executive is required to submit a detailed budget proposal for their portfolio.\n\n"
    "Please include all anticipated expenses for the upcoming year, even if you are unsure whether funding will ultimately be approved. The goal is to create an accurate picture of what your portfolio may require to operate successfully.\n\n"
    "For each expense, please:\n"
    "• Research and estimate the expected cost as accurately as possible.\n"
    "• Provide links, quotes, or rationale where applicable.\n"
    "• Explain what the expense is for.\n"
    "• Indicate whether the expense is a necessity or simply beneficial (5 being essential, 1 being optional). \n"
    "• Describe the impact on your portfolio if this expense is not funded.\n\n"
    "Please use your transition manuals, previous budgets, and past executive experiences as guidance when identifying anticipated expenses.\n\n"
    "Budget proposals should reflect what your portfolio would ideally need to achieve its goals for the year. Final allocations will be determined by the Executive Team based on overall organizational priorities and available funding.\n\n"
    "Please be thorough. It is easier to reduce a proposed budget than to identify missing expenses later in the year.   \n\n"
    "Please make a COPY of this whole file.... note along the bottom there is both the budget request and the portfolio summary, please do BOTH :) \n\n"
    "Thank you so much!"
)
HEADERS = ["Expense Item", "Category", "Estimated Cost ($)", "Explanation / Purpose",
           "Necessity Level", "Consequences if Not Funded", "Research / Quote Source"]


def main():
    wb = openpyxl.Workbook()

    # ---------------- budget request sheet ----------------
    ws = wb.active
    ws.title = "budget request"

    thin = Side(style="thin", color="C7CEDB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    money = '#,##0'

    # QIAA instructions block (A1:G12), as in the original template
    ws.cell(row=1, column=1, value=QIAA_INTRO)
    ws.merge_cells(start_row=1, start_column=1, end_row=12, end_column=7)
    intro = ws.cell(row=1, column=1)
    intro.alignment = Alignment(wrap_text=True, vertical="top")
    intro.font = Font(size=10)

    # header row 14
    for col, h in enumerate(HEADERS, start=1):
        ws.cell(row=14, column=col, value=h)
    hdr_fill = PatternFill("solid", fgColor=HDR_FILL)
    for col in range(1, 8):
        c = ws.cell(row=14, column=col)
        c.fill = hdr_fill
        c.font = Font(bold=True, color="000000", size=11)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[14].height = 30

    # context note row 13 (just above header)
    note = ("All figures in CAD. This is the full conference slate we're planning for 2026-27 (about $43,587 all-in). "
            "Delegate fees, dues, and fundraising bring in roughly $29,150 of that. The remaining ~$14,437 is "
            "what we'd run at a loss and are asking QIAA to cover. Registration is budgeted at regular-window "
            "rates to stay safe, so registering early would bring some of that back as a cushion.")
    ws.cell(row=13, column=1, value=note)
    ws.merge_cells(start_row=13, start_column=1, end_row=13, end_column=7)
    nc = ws.cell(row=13, column=1)
    nc.font = Font(italic=True, color=NAVY, size=9)
    nc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[13].height = 56

    # clear the example row + anything below
    for r in range(15, ws.max_row + 1):
        for col in range(1, 8):
            ws.cell(row=r, column=col).value = None

    all_rows = CONF_ROWS + TRAVEL_TAIL + OPERATING_ROWS
    r = 15
    for i, (item, cat, cost, expl, nec, cons, src) in enumerate(all_rows):
        zebra = ZEBRA if i % 2 else "FFFFFF"
        fill = PatternFill("solid", fgColor=zebra)
        vals = [item, cat, cost, expl, nec, cons, src]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.fill = fill
            c.border = border
            c.alignment = Alignment(wrap_text=True, vertical="top",
                                    horizontal="center" if col in (3, 5) else "left")
            c.font = Font(size=10)
            if col == 3:
                c.number_format = money
        ws.row_dimensions[r].height = 42
        r += 1

    total = sum(x[2] for x in all_rows)
    # TOTAL row
    tfill = PatternFill("solid", fgColor=TOTAL_FILL)
    ws.cell(row=r, column=1, value="TOTAL PROGRAM (Total Requested Budget)")
    ws.cell(row=r, column=3, value=total)
    for col in range(1, 8):
        c = ws.cell(row=r, column=col)
        c.fill = tfill
        c.font = Font(bold=True, color="000000", size=11)
        c.border = border
        c.alignment = Alignment(vertical="center",
                                horizontal="center" if col == 3 else "left")
        if col == 3:
            c.number_format = money
    ws.row_dimensions[r].height = 24
    total_row = r
    r += 2

    # funding-context block
    ws.cell(row=r, column=1, value="FUNDING CONTEXT (not expense rows): how the program is paid for")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    c = ws.cell(row=r, column=1)
    c.font = Font(bold=True, color=NAVY, size=10)
    c.fill = PatternFill("solid", fgColor=GREYBG)
    r += 1
    ws.cell(row=r, column=1, value="Total program expense (from above)")
    ws.cell(row=r, column=3, value=total).number_format = money
    for col in (1, 3):
        ws.cell(row=r, column=col).font = Font(size=10)
    r += 1
    running = total
    for label, amt, expl in FUNDING_CONTEXT:
        ws.cell(row=r, column=1, value=label)
        cc = ws.cell(row=r, column=3, value=amt)
        cc.number_format = '#,##0;(#,##0)'
        ws.cell(row=r, column=4, value=expl)
        for col in (1, 3, 4):
            ws.cell(row=r, column=col).font = Font(size=10)
            ws.cell(row=r, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        running += amt
        r += 1
    # net
    ws.cell(row=r, column=1, value="NET CLUB POSITION (planned operating loss for QIAA to backstop)")
    nc2 = ws.cell(row=r, column=3, value=running)
    nc2.number_format = '#,##0;(#,##0)'
    for col in range(1, 5):
        c = ws.cell(row=r, column=col)
        c.font = Font(bold=True, color=RED, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    # column widths
    widths = {"A": 30, "B": 16, "C": 14, "D": 52, "E": 10, "F": 40, "G": 46}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ---------------- portfolio summary sheet ----------------
    ps = wb.create_sheet("portfolio summary")
    ps_headers = ["Executive Name", "Portfolio", "Total Requested Budget", "Top 3 Funding Priorities"]
    for col, h in enumerate(ps_headers, start=1):
        ps.cell(row=1, column=col, value=h)
    for col in range(1, 5):
        h = ps.cell(row=1, column=col)
        h.fill = PatternFill("solid", fgColor=HDR_FILL)
        h.font = Font(bold=True, color="000000", size=11)
        h.alignment = Alignment(vertical="center", wrap_text=True)
    ps.cell(row=2, column=1, value="Jack Guillemette")
    ps.cell(row=2, column=2, value="Model United Nations (travel team)")
    ps.cell(row=2, column=3, value=total)
    ps.cell(row=2, column=3).number_format = '$#,##0'
    priorities = (
        "1) Our anchor conferences: McMUN, NCSC, and ChoMUN. These are the top-ranked ones we build the year around. "
        "2) The two Montreal trips (McMUN and ConMUN), which give us 32 delegate spots at the lowest cost per person since we bus instead of fly. "
        "3) Covering the gap, about $14,437, which is what's left after delegate fees, dues, and fundraising."
    )
    ps.cell(row=2, column=4, value=priorities)
    for col in range(1, 5):
        cc = ps.cell(row=2, column=col)
        cc.alignment = Alignment(wrap_text=True, vertical="top")
        cc.font = Font(size=10)
    ps.column_dimensions["A"].width = 20
    ps.column_dimensions["B"].width = 26
    ps.column_dimensions["C"].width = 18
    ps.column_dimensions["D"].width = 80
    ps.row_dimensions[2].height = 120

    wb.save(OUT)
    print(f"Saved: {OUT}")
    print(f"Expense rows: {len(all_rows)}  |  TOTAL PROGRAM: ${total:,}")
    print(f"Net club position (loss to backstop): ${running:,}")


if __name__ == "__main__":
    main()
